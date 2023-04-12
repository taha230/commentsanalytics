# -*- encoding: utf-8 -*-
"""
Copyright (c) 2020 - present
"""

from django import template
from django.contrib.auth.decorators import login_required
from django.contrib import admin
from django.http import HttpResponse, HttpResponseRedirect
from django.template import loader
from django.urls import reverse
from django.shortcuts import render, redirect
from termcolor import colored
from django.contrib.auth.models import User
from django.contrib.auth import password_validation
# from authentication import CustomUser
from .models import Request_Type, Bulk, Bulk_Status, Request_Type, Request_Run_Type, Request_Status, Request, Plan, Plan_Status, Payment_Status, Payment_Type, Payment, User_Log, User_Other_Fields, Seen_Status, Ticket, Ticket_Message, Ticket_Status, Empty_Status, User_Type, Post, Post_Paragraph, Post_Status, Request_Process_Status, Discount_Code
from apps.authentication.views import send_email_from_admin_to_client, send_email_from_client_to_admin
import datetime
import requests
import uuid
import json
from django.contrib.sites.shortcuts import get_current_site
import smtplib
import csv
from xlsxwriter.workbook import Workbook
import openpyxl
import time
from reportlab.pdfgen import canvas
import io
from django.http import FileResponse
from reportlab.lib.colors import pink, green, brown, white, black, red, gray
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.units import mm
# from reportlab.platypus import Image
from reportlab.lib.enums import TA_JUSTIFY
from reportlab.lib.utils import ImageReader
import pandas as pd
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.admin.views.decorators import staff_member_required
from django.http import JsonResponse
import gevent
from multiprocessing import Manager
import multiprocessing
from pymongo import MongoClient
from bson.objectid import ObjectId
from .forms import FileForm
import datetime


########################################### Global Variables #################################################

ROW_LIST_SHOW_COUNT = 10
MONTH_COUNT_REPORT = 12
WEEK_COUNT_REPORT = 7
POST_BLOG_COUNT_REPORT = 5
POST_HOMEPAGE_COUNT_REPORT = 3
BULK_SIZE_LIMIT = 15000
SPLIT_TOKEN = '###COMMENTSANALYTICS###'
IP_SINGLE_API = '138.201.111.134' # '95.217.33.222'
IP_SINGLE_API_CATEGORY = '167.235.207.111' #'136.243.77.239' # '167.235.207.111'

########################################### MongoDB Init ####################################################

mongo_address = 'localhost'
DB_NAME = 'CommentsAnalytics'
COLLECTION_NAME_BULKS = 'Bulks'
COLLECTION_NAME_REQUESTS = 'Requests'
COLLECTION_NAME_User_Log= 'User_Log'
client = MongoClient(mongo_address, 27017)
db = client[DB_NAME]
collection_Bulks = db[COLLECTION_NAME_BULKS]
collection_Requests = db[COLLECTION_NAME_REQUESTS]
collection_User_Log = db[COLLECTION_NAME_User_Log]

########################################### MongoDB Indexes #################################################

# collection_Bulks.create_index([ ("bulk_start_time_slot", -1) ])
# collection_Bulks.create_index([ ("user", -1) ])
# collection_Bulks.create_index([ ("status", -1) ])
# collection_Requests.create_index([ ("request_time_slot", -1) ])
# collection_Requests.create_index([ ("run_type", -1) ])
# collection_Requests.create_index([ ("status", -1) ])
# collection_Requests.create_index([ ("bulk", -1) ])
# collection_Requests.create_index([ ("user", -1) ])
# collection_User_Log.create_index([ ("time_slot", -1) ])

############################################# Pages ###########################################################

def homepage(request):
    html_template = loader.get_template('home/homepage.html')
    requests_types_tuple = list(Request_Type.choices())
    requests_types = list((i[1]) for i in requests_types_tuple)

    row_list_post, total_count_post = get_post_list_page(request, 0, POST_HOMEPAGE_COUNT_REPORT, Post_Status.PUBLISHED.value)

    context = {'segment': 'index',
                "requests_types": requests_types,
                "latest_posts": row_list_post
              }

    return HttpResponse(html_template.render(context, request))

def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:

        load_template = request.path.split('/')[-1]

        if load_template == 'admin':
            return HttpResponseRedirect(reverse('admin:index'))
        context['segment'] = load_template

        html_template = loader.get_template('home/' + load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:

        html_template = loader.get_template('home/page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:
        html_template = loader.get_template('home/page-500.html')
        return HttpResponse(html_template.render(context, request))

def sitemap(request):
    return render(request, 'home/sitemap.xml', content_type='text/xml')

def robots(request):
    return render(request, 'home/robots.txt', content_type='text/plain')

def requests_demo(request):

    requests_types_tuple = list(Request_Type.choices())
    requests_types = list((i[1]) for i in requests_types_tuple)
    row_list_post, total_count_post = get_post_list_page(request, 0, POST_HOMEPAGE_COUNT_REPORT, Post_Status.PUBLISHED.value)

    request_result = 'not exists'

    try:
        if request.method == "POST": 
            business_name = request.POST.get('input_business_name').strip()
            request_type = request.POST.get('input_request_type').strip()

            # insert demo request mongo
            request_result = insert_demo_request_db_mongodb(request, business_name, request_type)
            if ( request_result != 'not exists' ):
                request_result_length = len(request_result)
                one_eights_length = int(request_result_length/8)
                one_forth_length = int(request_result_length/4)
                if (one_forth_length >=1):
                    # request_result=  request_result[0:1 *one_third_length] +  ''.join('*' for x in range(5 * one_third_length)) + request_result[6 * one_third_length: 7 * one_third_length] + ''.join('*' for x in range( request_result_length - (7 * one_third_length)))
                    request_result=  request_result[0:1 *one_forth_length] +  ''.join('*' for x in range(2 * one_forth_length)) + request_result[3 * one_forth_length:]

    except Exception as e:
        print(e)
        pass
    
    return render(request, "home/homepage.html", {"msg": 'SUCCESS',
                                                                "segment": 'requests-demo', 
                                                                "requests_types": requests_types,
                                                                "request_result": request_result,
                                                                "latest_posts": row_list_post})

def homepage_pricing(request):

    html_template = loader.get_template('home/homepage-pricing.html')
    plan_list_active, total_enable = get_plan_list_page(request, 0, 100, Plan_Status.ENABLE)


    context = {'segment': 'index',
               "plan_list": plan_list_active
               ,'related_links' : [
               ]
               ,'related_posts' : [
                    {'url': '/Post/Free%20plan:%20Best%20approach%20to%20attract%20customers%20to%20purchase/', 'text': 'Free plan: Best approach to attract customers to purchase'}
                    ,{'url': '/Post/Bulk%20Requests%20needs%20for%20saving%20your%20time%20for%20large%20files/', 'text': 'Bulk Requests needs for saving your time for large files'}
                    ,{'url': '/Post/Single%20Request%20option%20to%20validate%20the%20quality%20of%20URL%20finding%20services/', 'text': 'Single Request option to validate the quality of URL finding services'}
                    ,{'url': '/Post/Bulk%20Requests%20Size%20Limit%20set%20to%2015%20K/', 'text': 'Bulk Requests Size Limit set to 15 K'}
                    ,{'url': '/Post/No%20expiration%20date%20for%20profile%20browse%20services/', 'text': 'No expiration date for profile browse services'}
               ]
              }

    return HttpResponse(html_template.render(context, request))

def homepage_contact(request):

    html_template = loader.get_template('home/homepage-contact.html')

    context = {'segment': 'index'
               ,'related_links' : [
               ]
               ,'related_posts' : [
                    {'url': '/Post/Free%20plan:%20Best%20approach%20to%20attract%20customers%20to%20purchase/', 'text': 'Free plan: Best approach to attract customers to purchase'}
                    ,{'url': '/Post/Bulk%20Requests%20needs%20for%20saving%20your%20time%20for%20large%20files/', 'text': 'Bulk Requests needs for saving your time for large files'}
                    ,{'url': '/Post/Single%20Request%20option%20to%20validate%20the%20quality%20of%20URL%20finding%20services/', 'text': 'Single Request option to validate the quality of URL finding services'}
                    ,{'url': '/Post/Bulk%20Requests%20Size%20Limit%20set%20to%2015%20K/', 'text': 'Bulk Requests Size Limit set to 15 K'}
                    ,{'url': '/Post/No%20expiration%20date%20for%20profile%20browse%20services/', 'text': 'No expiration date for profile browse services'}
               ]
              }

    return HttpResponse(html_template.render(context, request))

def homepage_overview(request):
    html_template = loader.get_template('home/homepage-overview.html')

    context = {'segment': 'index'
               ,'related_links' : []
               ,'related_posts' : []
              }

    return HttpResponse(html_template.render(context, request))

def homepage_blog(request):

    html_template = loader.get_template('home/homepage-blog.html')
    current_pagination = 1    
    last_pagination = 1 
    total_count_post = 0
    row_list_post = []

    try:

        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * POST_BLOG_COUNT_REPORT

        row_list_post, total_count_post = get_post_list_page(request, start, POST_BLOG_COUNT_REPORT, Post_Status.PUBLISHED.value)

        if (total_count_post == 0):
            last_pagination = 1
        elif (total_count_post % POST_BLOG_COUNT_REPORT == 0):
            last_pagination = int (total_count_post / POST_BLOG_COUNT_REPORT)
        else:
            last_pagination = int (total_count_post / POST_BLOG_COUNT_REPORT) + 1

    except Exception as e:
        print(e)
        pass
    
    context = {'segment': 'index',
                "current_pagination": current_pagination,
                "last_pagination": last_pagination,
                "page_limit": min(total_count_post, POST_BLOG_COUNT_REPORT),
                "total_items": total_count_post,
                "post_list": row_list_post
                ,'related_links' : [
                    {'url': '/sentiment-analysis/', 'text': 'Sentiment analysis tools to explore the customers\' emotions within the texts'}
                    ,{'url': '/key-phrases-extraction/', 'text': 'Use key phrase extraction to quickly identify the main concepts in text'}
                    ,{'url': '/named-entity-recognition/', 'text': 'Named Entity Recognition: identifying key information (entities) in comments'}
                    ,{'url': '/predict-customers-needs/', 'text': 'How to predict customers\' needs in their reviews'}
                ]
               ,'related_posts' : [
                    ]
              }

    return HttpResponse(html_template.render(context, request))

def homepage_blog_post_id(request):

    html_template = loader.get_template('home/homepage-blog-post.html')
    post_json = {}

    try:
        post_id = '0'
        try:
            post_id = request.path.split('/Post_')[-1][:-1]
        except Exception as e:
            pass
        
        post_json = get_post_details(post_id)

    except Exception as e:
        print(e)
        pass
    context = {'segment': 'index',
                          "post_json": post_json,
                          "title" : "profile browse | " + post_json['title'],
                          "description": post_json['title']
                ,'related_links' : [
               ]
              }

    return HttpResponse(html_template.render(context, request))

def homepage_blog_post_name(request):

    html_template = loader.get_template('home/homepage-blog-post.html')
    post_json = {}

    try:
        post_name = '0'
        try:
            post_name = request.path.split('/Post/')[-1][:-1]
        except Exception as e:
            pass
        post_json = get_post_details_by_name(post_name)

    except Exception as e:
        print(e)
        pass
    context = {'segment': 'index',
                          "post_json": post_json,
                          "title" : "CommentsAnalytics | " + post_json['title'],
                          "description": post_json['title']
                ,'related_links' : [
                     {'url': '/sentiment-analysis/', 'text': 'Sentiment analysis tools to explore the customers\' emotions within the texts'}
                    ,{'url': '/key-phrases-extraction/', 'text': 'Use key phrase extraction to quickly identify the main concepts in text'}
                    ,{'url': '/named-entity-recognition/', 'text': 'Named Entity Recognition: identifying key information (entities) in comments'}
                    ,{'url': '/predict-customers-needs/', 'text': 'How to predict customers\' needs in their reviews'}
               ]
              }

    return HttpResponse(html_template.render(context, request))

def homepage_services(request):
    html_template = loader.get_template('home/homepage-services.html')

    context = {'segment': 'index'
               ,'related_links' : [
               ]
               ,'related_posts' : [
                    {'url': '/Post/Business%20Profiles%20monitoring%20as%20a%20need%20for%20companies/', 'text': 'Business Profiles monitoring as a need for companies'}
                    ,{'url': '/Post/Zoominfo%20service%20is%20accesible/', 'text': 'Zoominfo service is accesible'}
                    ,{'url': '/Post/Crunchbase%20Profile%20URL%20finder%20added!/', 'text': 'Crunchbase Profile URL finder added'}
                    ,{'url': '/Post/Social%20network%20accounts%20extractor%20as%20a%20need%20for%20social%20marketers%20and%20growing%20your%20business/', 'text': 'Social network accounts extractor for social marketers and growing your business'}
               ]
              }

    return HttpResponse(html_template.render(context, request))

def homepage_terms(request):
    html_template = loader.get_template('home/homepage-terms.html')

    context = {'segment': 'index'
               ,'related_links' : [
               ]
               ,'related_posts' : [
                    ]
              }

    return HttpResponse(html_template.render(context, request))

def homepage_privacy(request):
    html_template = loader.get_template('home/homepage-privacy.html')

    context = {'segment': 'index'
               ,'related_links' : [
               ]
               ,'related_posts' : [
                    ]
              }

    return HttpResponse(html_template.render(context, request))

def homepage_chrome_extension(request):
    html_template = loader.get_template('home/homepage-chrome-extension.html')

    context = {'segment': 'index'
               ,'related_links' : [
               ]
               ,'related_posts' : [
                    ]
              }

    return HttpResponse(html_template.render(context, request))

############################################# Modules #########################################################

@login_required(login_url="/login/")
def get_user_list_page(request, start, limit):
    
    try:
        ################################ Select from auth_user Table #################################

        row_list = []

        # sorted by date_joined
        row_list_all = sorted(User.objects.all(), key = lambda x: x.date_joined, reverse = True)[start : start + limit]
        total_count = User.objects.all().count()

        for index, row in enumerate(list(row_list_all)):
            try:
                # if (index < start):
                #     continue
                # if (index >= (start+limit)):
                #     break

                row_json = {}
                row_json['index'] = index + 1 + start
                row_json['id'] = row.id
                row_json['remain_count'] = 0
                row_json['user_key'] = '-1'
                try:
                    user_other_fields_obj = User_Other_Fields.objects.filter(user_id=row.id)[0]

                    expired_date_short = user_other_fields_obj.expired_date
                    try:
                        expired_date_short = user_other_fields_obj.expired_date.strftime("%b %d, %Y")
                    except Exception as e:
                        print(e)
                    row_json['expired_date'] = expired_date_short
                    row_json['is_expired'] = True
                    
                    try:
                        if (user_other_fields_obj.expired_date.replace(tzinfo=None) > datetime.datetime.now().replace(tzinfo=None)):
                            row_json['is_expired'] = False
                    except Exception as e:
                        print(e)
                    
                    row_json['remain_count'] =  user_other_fields_obj.remain_count
                    row_json['user_key'] = user_other_fields_obj.user_key
                except Exception as e:
                    pass

                last_login_short = row.last_login
                try:
                    last_login_short = row.last_login.strftime("%b %d, %Y %H:%M")
                except Exception as e:
                    print(e)
                row_json['last_login'] = last_login_short

                date_joined_short = row.date_joined
                try:
                    date_joined_short = row.date_joined.strftime("%b %d, %Y")
                except Exception as e:
                    print(e)
                row_json['date_joined'] = date_joined_short

                row_json['is_superuser'] = row.is_superuser
                row_json['username'] = row.username
                row_json['password'] = row.password
                row_json['first_name'] = row.first_name
                row_json['last_name'] = row.last_name
                row_json['email'] = row.email
                row_json['is_active'] = row.is_active
                row_json['is_staff'] = row.is_staff

                user_plan_name = 'Admin'
                if (row.is_superuser == 0):
                    user_plan_name = get_user_plan_name(row.id)

                row_json['plan'] = user_plan_name

                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue

        print(colored('User select successfully from admin user !!!', 'green'))

        return row_list, total_count

    except Exception as e:
        print(colored('Unsuccessful User selection !!!', 'red'))
        print(e)
        return [], 0

def get_user_plan_name(user_id):
    plan_name = 'Admin'
    try:
        plan_name = User_Other_Fields.objects.all().filter(user_id= user_id)[0].plan
    except Exception as e:
        print(e)
    return plan_name

def diff_month(d1, d2):
    return (d1.year - d2.year) * 12 + d1.month - d2.month

def diff_day(d1, d2):
    
    return (d1.year - d2.year) * 365 + (d1.month - d2.month) * 30 + (d1.day - d2.day) 

def get_business_names_list_page(input_list, start, limit):

    try:
        ################################ Select from input Business Names #################################

        row_list = []
        total_count = len(input_list)

        for index, row in enumerate(input_list):
            try:

                if (index < start):
                    continue
                if (index >= (start+limit)):
                    break
                    
                row_json = {}
                row_json['index'] = index + 1
                row_json['name'] = row
                
                row_list.append(row_json) 

            except Exception as e:
                print(e)
                continue

        print(colored('business_names select successfully from user !!!', 'green'))

        return row_list, total_count

    except Exception as e:
        print(colored('Unsuccessful User selection !!!', 'red'))
        print(e)
        return [], 0

def get_bulk_list_page(request, start, limit, user):
    try:
        ################################ Select from home_Bulk Table #################################

        row_list = []
        total_count = 0
        # sorted by bulk_start_time_slot
        if (user == None): # return all bulks (Admin)
            row_list_all = sorted(Bulk.objects.all(), key = lambda x: x.bulk_start_time_slot, reverse = True)[start : start + limit]
            total_count = Bulk.objects.all().count()

        else: # filtered by user (Client)
            row_list_all = sorted(Bulk.objects.all().filter(user_id=request.user.id), key = lambda x: x.bulk_start_time_slot, reverse = True)[start : start + limit]
            total_count = Bulk.objects.all().filter(user_id=request.user.id).count()


        for index, row in enumerate(list(row_list_all)):
            try:

                # if (index < start):
                #     continue
                # if (index >= (start+limit)):
                #     break

                row_json = {}
                row_json['index'] = index + 1 + start
                row_json['user'] = ''
                try:
                    row_json['user'] = User.objects.all().filter(id=row.user_id)[0].email
                except Exception as e:
                    pass

                calculated_passed_count = row.passed_count
                
                row_json['title'] = row.name
                row_json['bulk_id'] = row.id
                row_json['request_type'] = row.request_type
                row_json['start_time'] = row.bulk_start_time_slot
                row_json['end_time'] = row.bulk_end_time_slot
                row_json['passed_count'] = calculated_passed_count
                row_json['total_count'] = row.total_count
                row_json['status'] = row.status
                row_json['percent'] = int(calculated_passed_count / row.total_count * 100)
                
            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue

        print(colored('Bulk select successfully from client user !!!', 'green'))

        return row_list, total_count

    except Exception as e:
        print(colored('Unsuccessful Bulk selection !!!', 'red'))
        print(e)
        return [], 0

def get_bulk_list_page_mongodb(request, start, limit, user, status=None , request_type= None):
    try:
        ################################ Select from home_Bulk Table #################################
        row_list = []
        total_count = 0
        # sorted by bulk_start_time_slot
        if (user == None): # return all bulks (Admin)
            query_bulk = {}
            if (status != None):
                query_bulk['status'] = status.name
            if (request_type != None):
                query_bulk['request_type'] = request_type.value
            row_list_all = collection_Bulks.find(query_bulk).sort([("bulk_start_time_slot", -1)])[start : start + limit]
            total_count = collection_Bulks.find(query_bulk).count()

        else: # filtered by user (Client)
            query_bulk = {'user': str(request.user.id)}
            if (status != None):
                query_bulk['status'] = status.name
            if (request_type != None):
                query_bulk['request_type'] = request_type.value
            row_list_all = collection_Bulks.find(query_bulk).sort([("bulk_start_time_slot", -1)])[start : start + limit]
            total_count = collection_Bulks.find(query_bulk).count()


        for index, row in enumerate(list(row_list_all)):
            try:
                row_json = {}
                row_json['index'] = index + 1 + start
                row_json['user'] = ''
                try:
                    row_json['user'] = User.objects.all().filter(id=row['user'])[0].email
                except Exception as e:
                    pass

                calculated_passed_count = row['passed_count']
                
                row_json['title'] = row['name']
                row_json['bulk_id'] = str(row['_id'])
                row_json['request_type'] = row['request_type']
                row_json['start_time'] = row['bulk_start_time_slot']
                row_json['end_time'] = row['bulk_end_time_slot']
                row_json['passed_count'] = calculated_passed_count
                row_json['total_count'] = row['total_count']
                row_json['status'] = row['status']
                row_json['percent'] = int(calculated_passed_count / row['total_count'] * 100)
                
                # calculate Positive/Neutral/Negetive count from 'Sentiment Analysis' Bulks
                if (row_json['request_type'] == Request_Type.SENTIMENT_ANALYSIS.value):
                    query_positive = {'bulk': row_json['bulk_id'], 'user': str(request.user.id), 'result': 'Positive'}
                    query_neutral = {'bulk': row_json['bulk_id'], 'user': str(request.user.id), 'result': 'Neutral'}
                    query_negative = {'bulk': row_json['bulk_id'], 'user': str(request.user.id), 'result': 'Negative'}
                    positive_count = collection_Requests.find(query_positive).count()
                    neutral_count = collection_Requests.find(query_neutral).count()
                    negative_count = collection_Requests.find(query_negative).count()
                    try:
                        row_json['positive_count'] = positive_count
                        row_json['neutral_count'] = neutral_count
                        row_json['negative_count'] = negative_count

                        row_json['positive_percent'] = int(positive_count / row['total_count'] * 100)
                        row_json['neutral_percent'] = int(neutral_count / row['total_count'] * 100)
                        row_json['negative_percent'] = int(negative_count / row['total_count'] * 100)

                    except Exception as e:
                        print(e)
                        row_json['positive_count'] = 0
                        row_json['neutral_count'] = 0
                        row_json['negative_count'] = 0

                        row_json['positive_percent'] = 0
                        row_json['neutral_percent'] = 0
                        row_json['negative_percent'] = 0


                
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue

        print(colored('Bulk select successfully from client user !!!', 'green'))

        return row_list, total_count

    except Exception as e:
        print(colored('Unsuccessful Bulk selection !!!', 'red'))
        print(e)
        return [], 0

def get_request_list_page(request, start, limit, user, run_type):

    try:
        ################################ Select from home_Request Table #################################

        row_list = []
        total_count = 0

        # sorted by request_time_slot
        if (user == None): # return all bulks (Admin)
            row_list_all = sorted(Request.objects.all().filter(run_type= run_type.value), key = lambda x: x.request_time_slot, reverse = True)[start : start + limit]
            total_count = Request.objects.all().filter(run_type= run_type.value).count()

        else: # filtered by user (Client)
            row_list_all = sorted(Request.objects.all().filter(user_id=request.user.id, run_type= run_type.value), key = lambda x: x.request_time_slot, reverse = True)[start : start + limit]
            total_count = Request.objects.all().filter(user_id=request.user.id, run_type= run_type.value).count()


        for index, row in enumerate(list(row_list_all)):
            try:

                # if (index < start):
                #     continue
                # if (index >= (start+limit)):
                #     break

                row_json = {}
                row_json['index'] = index + 1 + start
                row_json['user'] = ''
                try:
                    row_json['user'] = User.objects.all().filter(id=row.user_id)[0].email
                except Exception as e:
                    pass
                
                row_json['business_name'] = row.query
                row_json['request_type'] = row.request_type
                row_json['request_time'] = row.request_time_slot
                row_json['result_time'] = row.result_time_slot
                row_json['result'] = row.result
                row_json['status'] = row.status
                
            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('Request select successfully from client user !!!', 'green'))

        return row_list, total_count

    except Exception as e:
        print(colored('Unsuccessful Request selection !!!', 'red'))
        print(e)
        return [], 0

def get_request_list_page_mongodb(request, start, limit, user, run_type, request_type= None):

    try:
        ################################ Select from home_Request Table #################################

        row_list = []
        total_count = 0

        # sorted by request_time_slot
        if (user == None): # return all bulks (Admin)
            query_request = {'run_type': run_type.value}
            if (request_type != None):
                query_request['request_type'] = request_type
            row_list_all = collection_Requests.find(query_request).sort([("request_time_slot", -1)])[start : start + limit]
            total_count = collection_Requests.find(query_request).count()
        else: # filtered by user (Client)
            query_request = {'user': str(request.user.id), 'run_type': run_type.value}
            if (request_type != None):
                query_request['request_type'] = request_type
            row_list_all = collection_Requests.find(query_request).sort([("request_time_slot", -1)])[start : start + limit]
            total_count = collection_Requests.find(query_request).count()



        for index, row in enumerate(list(row_list_all)):
            try:
                row_json = {}
                row_json['index'] = index + 1 + start
                row_json['user'] = ''
                try:
                    row_json['user'] = User.objects.all().filter(id=row['user'])[0].email
                except Exception as e:
                    pass
                
                row_json['text_query'] = row['query']
                row_json['request_type'] = row['request_type']
                row_json['request_time'] = row['request_time_slot']
                row_json['result_time'] = row['result_time_slot']
                row_json['result'] = row['result']
                row_json['status'] = row['status']
                
            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('Request select successfully from client user !!!', 'green'))

        return row_list, total_count

    except Exception as e:
        print(colored('Unsuccessful Request selection !!!', 'red'))
        print(e)
        return [], 0

def get_request_list_bulk_export(request, bulk_id):

    try:
        ################################ Select from home_Request Table #################################

        row_list = []
        total_count = 0

        # sorted by order
        row_list_all = sorted(Request.objects.all().filter( bulk_id = bulk_id), key = lambda x: x.order)


        for index, row in enumerate(list(row_list_all)):
            try:
                row_json = {}

                row_json['business_name'] = row.query
                row_json['request_type'] = row.request_type
                row_json['request_time'] = row.request_time_slot
                row_json['result_time'] = row.result_time_slot
                row_json['result'] = row.result
                
            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('Request Bulk for Download select successfully from client user !!!', 'green'))

        return row_list

    except Exception as e:
        print(colored('Unsuccessful Request Bulk for Download selection !!!', 'red'))
        print(e)
        return [], 0

def get_request_list_bulk_export_mongodb(request, bulk_id):

    try:
        ################################ Select from Request Table Mongo #################################

        row_list = []
        total_count = 0

        query_bulk_id = {'bulk' : bulk_id}
        row_list_all = collection_Requests.find(query_bulk_id)


        for index, row in enumerate(list(row_list_all)):
            try:
                row_json = {}

                row_json['business_name'] = row['query']
                row_json['request_type'] = row['request_type']
                row_json['request_time'] = row['request_time_slot']
                row_json['result_time'] = row['result_time_slot']
                row_json['result'] = row['result']
                if ('result_sentiment' in row):
                    row_json['result_sentiment'] = row['result_sentiment']

                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('Request Bulk for Download select successfully from client user !!!', 'green'))

        return row_list

    except Exception as e:
        print(colored('Unsuccessful Request Bulk for Download selection !!!', 'red'))
        print(e)
        return [], 0


def get_request_list_bulk_mongodb(request, bulk_id):

    try:
        ################################ Select from Request Table Mongo #################################

        row_list = []
        total_count = 0

        query_bulk_id = {'bulk' : bulk_id}
        row_list_all = collection_Requests.find(query_bulk_id)


        for index, row in enumerate(list(row_list_all)):
            try:
                row_json = {}
                row_json['index'] = index + 1
                row_json['text_query'] = row['query']
                row_json['request_type'] = row['request_type']
                row_json['request_time'] = row['request_time_slot']
                row_json['result_time'] = row['result_time_slot']
                row_json['result'] = row['result']
                if ('result_sentiment' in row):
                    row_json['sentiment'] = row['result_sentiment']
                
            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('Request Bulk for Download select successfully from client user !!!', 'green'))

        return row_list

    except Exception as e:
        print(colored('Unsuccessful Request Bulk for Download selection !!!', 'red'))
        print(e)
        return [], 0

def get_user_log_list_page(request, start, limit, user):

    try:
        ################################ Select from home_user_log Table #################################

        row_list = []

        # sorted by request_time_slot
        if (user == None): # return all bulks (Admin)
            row_list_all = sorted(User_Log.objects.all(), key = lambda x: x.time_slot, reverse = True)[start : start + limit]
            total_count = User_Log.objects.all().count()
        else: # filtered by user (Client)
            row_list_all = sorted(User_Log.objects.all().filter(user_id=request.user.id), key = lambda x: x.time_slot, reverse = True)[start : start + limit]
            total_count = User_Log.objects.all().filter(user_id=request.user.id).count()

        for index, row in enumerate(list(row_list_all)):
            try:

                # if (index < start):
                #     continue
                # if (index >= (start+limit)):
                #     break

                row_json = {}
                row_json['index'] = index + 1 + start
                row_json['user'] = ''
                row_json['user_name'] = ''
                try:
                    row_json['user'] = User.objects.all().filter(id=row.user_id)[0].email
                    row_json['username'] = User.objects.all().filter(id=row.user_id)[0].username
                except Exception as e:
                    pass
                
                row_json['activity'] = row.activity
                row_json['time_slot'] = row.time_slot

                user_plan_name = 'Admin'
                if (row.user.is_superuser == 0):
                    user_plan_name = get_user_plan_name(row.user_id)

                row_json['plan'] = user_plan_name
                
            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('User_Log select successfully from client user !!!', 'green'))

        return row_list, total_count

    except Exception as e:
        print(colored('Unsuccessful User_Log selection !!!', 'red'))
        print(e)
        return [], 0

def get_user_log_list_page_mongodb(request, start, limit, user):

    try:
        ################################ Select from user_log Table Mongodb #################################

        row_list = []

        # sorted by request_time_slot
        if (user == None): # return all bulks (Admin)
            query_log = {}
            row_list_all = collection_User_Log.find(query_log).sort([("time_slot", -1)])[start : start + limit]
            total_count = collection_User_Log.find(query_log).count()

        else: # filtered by user (Client)
            query_log = {'user': request.user.id}
            row_list_all = collection_User_Log.find(query_log).sort([("time_slot", -1)])[start : start + limit]
            total_count = collection_User_Log.find(query_log).count()


        for index, row in enumerate(list(row_list_all)):
            try:

        
                row_json = {}
                row_json['index'] = index + 1 + start
                row_json['user'] = ''
                row_json['user_name'] = ''
                try:
                    row_json['user'] = User.objects.all().filter(id=row['user'])[0].email
                    row_json['username'] = User.objects.all().filter(id=row['user'])[0].username
                except Exception as e:
                    pass
                
                row_json['activity'] = row['activity']
                row_json['time_slot'] = row['time_slot']

                user_plan_name = 'Admin'
                if (row['user'] != "18"): # admin user id = 18
                    user_plan_name = get_user_plan_name(row['user'])

                row_json['plan'] = user_plan_name
                
            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('User_Log select successfully from client user !!!', 'green'))

        return row_list, total_count

    except Exception as e:
        print(colored('Unsuccessful User_Log selection !!!', 'red'))
        print(e)
        return [], 0

def get_month_list_request (request, run_type):

    monthly_request_list_value = [0] * MONTH_COUNT_REPORT # 12 months of the year
    monthly_request_list_month = ["Jan"] * MONTH_COUNT_REPORT # 12 months of the year

    monthly_request_list_json = { "data" :monthly_request_list_value, "months": monthly_request_list_month}


    try:
        ################################ Select from home_Request Table #################################

        row_list = []
        
        now = datetime.datetime.now()

        # get list of last 12 months
        month_list_recent = [now.strftime("%B")[:3]]
        for _ in range(1, MONTH_COUNT_REPORT):
            now = now.replace(day=1) - datetime.timedelta(days=1)
            month_list_recent.append(str(now.strftime("%B")[:3]))
        
        monthly_request_list_json['months'] = list(reversed(month_list_recent))

        time_past = datetime.datetime.now() - datetime.timedelta(days=365)
        
        # sorted by request_time_slot
        if (request == None): # return all Request (Admin)
            if (run_type == None):
                # row_list_all = sorted(Request.objects.all().filter(request_time_slot__gte = time_past, request_time_slot__lt = datetime.datetime.now()), key = lambda x: x.request_time_slot, reverse = True)
                query_request= {'request_time_slot': {'$lte': datetime.datetime.now()}, 'request_time_slot': {'$gte': time_past}}
                row_list_all = collection_Requests.find(query_request).sort([("request_time_slot", -1)])
            else:
                query_request= {'tun_type': run_type.value, 'request_time_slot': {'$lte': datetime.datetime.now()}, 'request_time_slot': {'$gte': time_past}}
                row_list_all = collection_Requests.find(query_request).sort([("request_time_slot", -1)])
        else: # filtered by user (Client)
            if (run_type == None):
                query_request= {'user': str(request.user.id), 'request_time_slot': {'$lte': datetime.datetime.now()}, 'request_time_slot': {'$gte': time_past}}
                row_list_all = collection_Requests.find(query_request).sort([("request_time_slot", -1)])
            else:
                query_request= {'run_type': run_type.value, 'user': str(request.user.id) ,'request_time_slot': {'$lte': datetime.datetime.now()}, 'request_time_slot': {'$gte': time_past}}
                row_list_all = collection_Requests.find(query_request).sort([("request_time_slot", -1)])
        
        total_count = row_list_all.count()


        for index, row in enumerate(list(row_list_all)):
            try:

                diff_month_value = diff_month(datetime.datetime.now(), row['request_time_slot'])
                if (diff_month_value < MONTH_COUNT_REPORT):
                    monthly_request_list_json['data'][MONTH_COUNT_REPORT-diff_month_value-1] += 1


            except Exception as e:
                print(e)
                continue


        print(colored('Request select successfully !!!', 'green'))

        return monthly_request_list_json

    except Exception as e:
        print(colored('Unsuccessful Request selection !!!', 'red'))
        print(e)
        return monthly_request_list_json

def get_week_list_request (request, run_type):

    weekly_request_list_value = [0] * WEEK_COUNT_REPORT # 7 day of the week
    weekly_request_list_week = ["SAT"] * WEEK_COUNT_REPORT # 7 day of the week

    weekly_request_list_json = { "data" :weekly_request_list_value, "months": weekly_request_list_week}


    try:
        ################################ Select from home_Request Table #################################

        row_list = []
        
        now = datetime.datetime.now()

        # get list of last 7 days
        week_list_recent = [now.strftime("%A")[:3]]
        for _ in range(1, WEEK_COUNT_REPORT):
            now = now - datetime.timedelta(days=1)
            week_list_recent.append(str(now.strftime("%A")[:3]))
        
        weekly_request_list_json['months'] = list(reversed(week_list_recent))

        time_past = datetime.datetime.now() - datetime.timedelta(days=8)

        
        # sorted by request_time_slot
        if (request == None): # return all Request (Admin)
            if (run_type == None):
                # row_list_all = sorted(Request.objects.all().filter(request_time_slot__gte = time_past, request_time_slot__lt = datetime.datetime.now()), key = lambda x: x.request_time_slot, reverse = True)
                query_request= {'request_time_slot': {'$lte': datetime.datetime.now()}, 'request_time_slot': {'$gte': time_past}}
                row_list_all = collection_Requests.find(query_request).sort([("request_time_slot", -1)])
            else:
                query_request= {'tun_type': run_type.value, 'request_time_slot': {'$lte': datetime.datetime.now()}, 'request_time_slot': {'$gte': time_past}}
                row_list_all = collection_Requests.find(query_request).sort([("request_time_slot", -1)])
        else: # filtered by user (Client)
            if (run_type == None):
                query_request= {'user': str(request.user.id), 'request_time_slot': {'$lte': datetime.datetime.now()}, 'request_time_slot': {'$gte': time_past}}
                row_list_all = collection_Requests.find(query_request).sort([("request_time_slot", -1)])
            else:
                query_request= {'run_type': run_type.value, 'user': str(request.user.id) ,'request_time_slot': {'$lte': datetime.datetime.now()}, 'request_time_slot': {'$gte': time_past}}
                row_list_all = collection_Requests.find(query_request).sort([("request_time_slot", -1)])
        total_count = row_list_all.count()


        for index, row in enumerate(list(row_list_all)):
            try:

                diff_day_value = diff_day(datetime.datetime.now(), row['request_time_slot'])
                if (diff_day_value < WEEK_COUNT_REPORT):
                    weekly_request_list_json['data'][WEEK_COUNT_REPORT-diff_day_value-1] += 1


            except Exception as e:
                print(e)
                continue


        print(colored('Request select successfully !!!', 'green'))
        return weekly_request_list_json

    except Exception as e:
        print(colored('Unsuccessful Request selection !!!', 'red'))
        print(e)
        return weekly_request_list_json

def get_transaction_list_page(request, start, limit, user, payment_status):

    total_earning = 0
    monthly_avg_earning = 0

    monthly_earning_list_value = [0] * MONTH_COUNT_REPORT # 12 months of the year
    monthly_earning_list_month = ["Jan"] * MONTH_COUNT_REPORT # 12 months of the year

    monthly_earning_list_json = { "data" :monthly_earning_list_value, "months": monthly_earning_list_month}

    try:
        ################################ Select from home_Payment Table #################################

        row_list = []
        
        now = datetime.datetime.now()

        # get list of last 12 months
        month_list_recent = [now.strftime("%B")[:3]]
        for _ in range(1, MONTH_COUNT_REPORT):
            now = now.replace(day=1) - datetime.timedelta(days=1)
            month_list_recent.append(str(now.strftime("%B")[:3]))
        
        monthly_earning_list_json['months'] = list(reversed(month_list_recent))


        
        # sorted by payment_time_slot
        if (user == None): # return all payments (Admin)
            row_list_all = sorted(Payment.objects.all().filter(status= payment_status.value), key = lambda x: x.payment_time_slot, reverse = True)
        else: # filtered by user (Client)
            row_list_all = sorted(Payment.objects.all().filter(user_id=request.user.id, status= payment_status.value), key = lambda x: x.payment_time_slot, reverse = True)
            
        total_count = len(row_list_all)


        for index, row in enumerate(list(row_list_all)):
            try:
                total_earning += row.price

                diff_month_value = diff_month(datetime.datetime.now(), row.payment_time_slot)
                if (diff_month_value < MONTH_COUNT_REPORT):
                    monthly_earning_list_json['data'][MONTH_COUNT_REPORT-diff_month_value-1] += row.price

                if (index < start):
                    continue
                if (index >= (start+limit)):
                    break

                row_json = {}
                row_json['index'] = index + 1
                row_json['user'] = ''
                try:
                    row_json['user'] = User.objects.all().filter(id=row.user_id)[0].email
                except Exception as e:
                    pass
                
                row_json['plan'] = row.plan
                row_json['id'] = row.id
                row_json['price'] = row.price
                row_json['count'] = row.plan.count
                row_json['payment_time'] = row.payment_time_slot
                row_json['type'] = row.payment_type
                row_json['status'] = row.status
                
            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('Payment select successfully !!!', 'green'))

        count_non_zero_earning_month = 0
        for month_earning in monthly_earning_list_json['data']:
            if (month_earning != 0):
                count_non_zero_earning_month += 1

        if (count_non_zero_earning_month != 0):
            monthly_avg_earning = total_earning / count_non_zero_earning_month
        else:
            monthly_avg_earning =0

        return row_list, total_count, total_earning, monthly_avg_earning, monthly_earning_list_json

    except Exception as e:
        print(colored('Unsuccessful Payment selection !!!', 'red'))
        print(e)
        return [], 0 , total_earning, monthly_avg_earning , monthly_earning_list_json

def get_ticket_list_page(request, start, limit, user):

    try:
        ################################ Select from home_Ticket Table #################################

        row_list = []
        total_count = 0
        unread_count = 0

        # sorted by request_time_slot
        if (user == None): # return all ticket (Admin)
            row_list_all = sorted(Ticket.objects.all(), key = lambda x: x.question_time_slot, reverse = True)[start : start + limit]
            total_count = Ticket.objects.all().count()
            unread_count = Ticket.objects.all().filter( admin_seen_status=Seen_Status.UNSEEN.value).count()
            

        else: # filtered by user (Client)
            row_list_all = sorted(Ticket.objects.all().filter(user_id=request.user.id), key = lambda x: x.answer_time_slot, reverse = True)[start : start + limit]
            total_count = Ticket.objects.all().filter(user_id=request.user.id).count()
            unread_count = Ticket.objects.all().filter(user_id=request.user.id, client_seen_status=Seen_Status.UNSEEN.value).count()



        for index, row in enumerate(list(row_list_all)):
            try:

                row_json = {}
                row_json['index'] = index + 1 + start
                row_json['ticket_id'] = row.id
                row_json['user'] = ''
                try:
                    row_json['user'] = User.objects.all().filter(id=row.user_id)[0].email
                except Exception as e:
                    pass
                
                row_json['title'] = row.title
                row_json['question_time_slot'] = row.question_time_slot
                row_json['answer_time_slot'] = row.answer_time_slot
                row_json['client_seen_status'] = row.client_seen_status
                row_json['admin_seen_status'] = row.admin_seen_status
                row_json['status'] = row.status
                row_json['empty_status'] = row.empty_status
                row_json['message_count'] = row.message_count
                
            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('Ticket select successfully from client user !!!', 'green'))

        return row_list, total_count, unread_count

    except Exception as e:
        print(colored('Unsuccessful Ticket selection !!!', 'red'))
        print(e)
        return [], 0, 0

def get_post_list_page(request, start, limit, post_status):

    try:
        ################################ Select from home_Post Table #################################

        row_list = []
        total_count = 0

        # post_status None for all status in blog admin
        if (post_status == None):
            total_count = Post.objects.all().count()
            # sorted by time_slot
            row_list_all = sorted(Post.objects.all(), key = lambda x: x.time_slot, reverse = True)[start : start + limit]
        else:
            total_count = Post.objects.all().filter(status = post_status).count()

            # sorted by time_slot
            row_list_all = sorted(Post.objects.all().filter(status = post_status), key = lambda x: x.time_slot, reverse = True)[start : start + limit]


        for index, row in enumerate(list(row_list_all)):
            try:

                row_json = {}
                row_json['index'] = index + 1 + start
                row_json['post_id'] = row.id
                row_json['title'] = row.title
                row_json['time_slot'] = row.time_slot
                row_json['image_name'] = row.image_name
                row_json['status'] = row.status

                if (post_status == None):
                    row_json['paragraph_list'] = []
                else:
                    post_paragraph_list, total_count_paragraph = get_post_paragraph_list(row.id)
                    row_json['paragraph_list'] = post_paragraph_list


            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('Post select successfully from client user !!!', 'green'))
        return row_list, total_count

    except Exception as e:
        print(colored('Unsuccessful Post selection !!!', 'red'))
        print(e)
        return [], 0

def get_discount_list_page(request, start, limit, discount_status):

    try:
        ################################ Select from home_Discount Table #################################

        row_list = []
        total_count = 0

        # discount_status None for all status in blog admin
        if (discount_status == None):
            total_count = Discount_Code.objects.all().count()
            # sorted by time_slot
            row_list_all = sorted(Discount_Code.objects.all(), key = lambda x: x.time_slot, reverse = True)[start : start + limit]
        else:
            total_count = Discount_Code.objects.all().filter(status = discount_status).count()

            # sorted by time_slot
            row_list_all = sorted(Discount_Code.objects.all().filter(status = discount_status), key = lambda x: x.time_slot, reverse = True)[start : start + limit]


        for index, row in enumerate(list(row_list_all)):
            try:

                row_json = {}
                row_json['index'] = index + 1 + start
                row_json['discount_id'] = row.id
                row_json['value'] = row.value
                row_json['time_slot'] = row.time_slot
                row_json['code'] = row.code
                row_json['status'] = row.status
            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('Discount select successfully from admin !!!', 'green'))
        return row_list, total_count

    except Exception as e:
        print(colored('Unsuccessful Discount selection !!!', 'red'))
        print(e)
        return [], 0

def get_post_details(post_id):

    try:
        ################################ Select from home_Post Table #################################

        row_json = {}

        if (post_id == '0'):
            return row_json
        else:
            # post_selected
            post_selected_list = Post.objects.all().filter(id = post_id)
            if (post_selected_list == None or post_selected_list == []):
                return row_json
            
            post_selected = post_selected_list[0]

            try:

                row_json['post_id'] = post_selected.id
                row_json['title'] = post_selected.title
                row_json['time_slot'] = post_selected.time_slot
                row_json['image_name'] = post_selected.image_name
                row_json['status'] = post_selected.status

                post_paragraph_list, total_count_paragraph = get_post_paragraph_list(post_selected.id)
                row_json['paragraph_list'] = post_paragraph_list

            except Exception as e:
                print(e)


        print(colored('Post select successfully from client user !!!', 'green'))
        return row_json

    except Exception as e:
        print(colored('Unsuccessful Post selection !!!', 'red'))
        print(e)
        return {}

def get_post_details_by_name(post_name):

    try:
        ################################ Select from home_Post Table #################################

        row_json = {}

        if (post_name == '0'):
            return row_json
        else:
            # post_selected
            print(colored(post_name, 'blue'))

            name_regexp = post_name + ''
            post_selected_list = Post.objects.all().filter(title=name_regexp)
            if (post_selected_list == None or post_selected_list == []):
                return {'title':'Not Found'}

            post_selected = post_selected_list[0]

            try:

                row_json['post_id'] = post_selected.id
                row_json['title'] = post_selected.title
                row_json['time_slot'] = post_selected.time_slot
                row_json['image_name'] = post_selected.image_name
                row_json['status'] = post_selected.status

                post_paragraph_list, total_count_paragraph = get_post_paragraph_list(post_selected.id)
                row_json['paragraph_list'] = post_paragraph_list

            except Exception as e:
                print(e)


        print(colored('Post select successfully from client user !!!', 'green'))
        return row_json

    except Exception as e:
        print(colored('Unsuccessful Post selection !!!', 'red'))
        print(e)
        return {'title':'Not Found'}

def get_ticket_messages_list(ticket_id):

    try:
        ################################ Select from home_Ticket_Message Table #################################

        row_list = []
        total_count = 0

        row_list_all = sorted(Ticket_Message.objects.all().filter(ticket_id=ticket_id), key = lambda x: x.time_slot, reverse = False)
        total_count = Ticket_Message.objects.all().filter(ticket_id=ticket_id).count()


        for index, row in enumerate(list(row_list_all)):
            try:

                row_json = {}
                row_json['index'] = index + 1
                row_json['message_id'] = row.id
                row_json['user'] = ''
                try:
                    row_json['user'] = User.objects.all().filter(id=row.user_id)[0].email
                except Exception as e:
                    pass
                
                row_json['body'] = row.body
                row_json['body_list'] = row.body.split('\n')
                row_json['seen_status'] = row.seen_status
                row_json['time_slot'] = row.time_slot
                row_json['user_type'] = row.user_type
                
            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('Ticket Message select successfully from client user !!!', 'green'))

        return row_list, total_count

    except Exception as e:
        print(colored('Unsuccessful Ticket Message selection !!!', 'red'))
        print(e)
        return [], 0

def get_post_paragraph_list(post_id):

    try:
        ################################ Select from home_Post_Paragraph Table #################################

        row_list = []
        total_count = 0

        row_list_all = sorted(Post_Paragraph.objects.all().filter(post_id=post_id), key = lambda x: x.time_slot, reverse = False)
        total_count = Post_Paragraph.objects.all().filter(post_id=post_id).count()


        for index, row in enumerate(list(row_list_all)):
            try:

                row_json = {}
                row_json['index'] = index + 1
                row_json['paragraph_id'] = row.id
                row_json['body'] = row.body
                row_json['body_list'] = row.body.split('\n')
                row_json['time_slot'] = row.time_slot
                
            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('Paragraph select successfully from client user !!!', 'green'))

        return row_list, total_count

    except Exception as e:
        print(colored('Unsuccessful Paragraph selection !!!', 'red'))
        print(e)
        return [], 0

def get_plan_list_page(request, start, limit, status):
    try:
        ################################ Select from home_plan Table #################################

        row_list = []

        # sorted by order
        if (status == None): # return all plans (Admin)
            row_list_all = sorted(Plan.objects.all(), key = lambda x: x.order)
        else: # filtered by status (Client)
            row_list_all = sorted(Plan.objects.all().filter(status= status.value), key = lambda x: x.order)
            
        total_count = len(row_list_all)
        for index, row in enumerate(list(row_list_all)):
            try:

                if (index < start):
                    continue
                if (index >= (start+limit)):
                    break

                row_json = {}
                row_json['index'] = index + 1
                row_json['name'] = row.name
                row_json['description'] = row.description
                row_json['has_list']=[]
                row_json['hosnot_list'] = []
                try:
                    description_json = json.loads(row.description)
                    if('has_list' in description_json):
                        row_json['has_list'] = description_json['has_list']
                    if('hasnot_list' in description_json):
                        row_json['hasnot_list']= description_json['hasnot_list']

                except Exception as e:
                    print('Exception in reading description : ' + str(e))

                row_json['count'] = row.count
                row_json['price'] = row.price
                row_json['discount'] = row.discount
                row_json['price_discounted'] = float(row.price * float(( 100 - row.discount )/100))
                row_json['status'] = row.status
                row_json['order'] = row.order
                row_json['tag'] = row.tag
                row_json['modify_time_slot'] = row.modify_time_slot
                row_json['plan_id'] = row.id
                
            
                row_list.append(row_json)

            except Exception as e:
                print(e)
                continue


        print(colored('Plan select successfully from admin user !!!', 'green'))

        return row_list, total_count

    except Exception as e:
        print(colored('Failed Plan selection !!!', 'red'))
        print(e)
        return [], 0

def insert_bulk_requests_list_multiprocess(request, bulk_obj, request_type, input_list):

    for query in input_list:
            while True:
                try:
                    request_obj = Request( request_time_slot = datetime.datetime.now(), result_time_slot=None, query = query, result=None, status = Request_Status.IN_QUEUE.value, request_type = request_type, run_type= Request_Run_Type.BULK.value, bulk=bulk_obj, user= request.user)
                    # request_obj = Request( request_time_slot = datetime.datetime.now(), result_time_slot=datetime.datetime.now(), query = query, result='https:/microsoft.com', status = Request_Status.SUCCESS.value, request_type = request_type, run_type= Request_Run_Type.BULK.value, bulk=bulk_obj, user= request.user, order = order)
                    request_obj.save()
                    
                    # inserted_list_verify = Request.objects.all().filter(id= request_obj.id)

                    # if len(inserted_list_verify) == 1:
                    #     break
                    break
                except Exception as e:
                    print(e)
                    continue

def insert_bulk_requests_list_multiprocess_mongodb(request, bulk_id, request_type, input_list):

    insert_list = []

    for index, query in enumerate(input_list):
        request_insert_json = {}
        request_insert_json['request_time_slot'] = datetime.datetime.now()
        request_insert_json['result_time_slot'] = None
        request_insert_json['query'] = query
        request_insert_json['result'] =  None
        request_insert_json['status'] = Request_Status.IN_QUEUE.value
        request_insert_json['request_type'] = request_type
        request_insert_json['run_type'] = Request_Run_Type.BULK.value
        request_insert_json['bulk'] = str(bulk_id)
        request_insert_json['user'] = str(request.user.id)
        
        insert_list.append(request_insert_json)

        if (len(insert_list) == 1000 or index == len(input_list)-1):
            while True:
                try:
                    collection_Requests.insert_many(insert_list)
                    insert_list = []
                    break    
                except Exception as e:
                    print(e)
                    continue

def insert_bulk_request_db (request, bulk_title, request_type, business_names_list):
    
    try:
        bulk_obj = Bulk( bulk_start_time_slot = datetime.datetime.now(), bulk_end_time_slot=None, name = bulk_title, status = Bulk_Status.IN_QUEUE.value, request_type = request_type, total_count= len(business_names_list), passed_count = 0, user= request.user)
        # bulk_obj = Bulk( bulk_start_time_slot = datetime.datetime.now(), bulk_end_time_slot=datetime.datetime.now(), name = bulk_title, status = Bulk_Status.COMPLETED.value, request_type = request_type, total_count= len(business_names_list), passed_count = len(business_names_list), user= request.user)
        bulk_obj.save()

        process_num = 30
        chunks = chunkIt(business_names_list, process_num)
        
        processes = []

        for i in range(len(chunks)):
            processes.append(multiprocessing.Process(target=insert_bulk_requests_list_multiprocess, args=[request, bulk_obj, request_type, chunks[i]]))

        for p in processes:
            p.start()

        for p in processes:
            p.join()

        order = 1
        # for query in business_names_list:
        #     while True:
        #         try:
        #             request_obj = Request( request_time_slot = datetime.datetime.now(), result_time_slot=None, query = query, result=None, status = Request_Status.IN_QUEUE.value, request_type = request_type, run_type= Request_Run_Type.BULK.value, bulk=bulk_obj, user= request.user)
        #             # request_obj = Request( request_time_slot = datetime.datetime.now(), result_time_slot=datetime.datetime.now(), query = query, result='https:/microsoft.com', status = Request_Status.SUCCESS.value, request_type = request_type, run_type= Request_Run_Type.BULK.value, bulk=bulk_obj, user= request.user, order = order)
        #             request_obj.save()
                    
        #             inserted_list_verify = Request.objects.all().filter(id= request_obj.id)

        #             if len(inserted_list_verify) == 1:
        #                 order += 1
        #                 break
        #         except Exception as e:
        #             print(e)
        #             continue
        
        requests_count = Request.objects.all().filter(bulk_id= bulk_obj.id).count()


        # update remain in User_Other_Fields in db
        update_remain_count_user(request, -1 * len(business_names_list))

    except Exception as e:
        print(e)
        pass

def insert_bulk_request_db_mongodb (request, bulk_title, request_type, business_names_list):
    
    try:
        bulk_insert_json = {}
        bulk_insert_json['bulk_start_time_slot'] = datetime.datetime.now()
        bulk_insert_json['bulk_end_time_slot'] = None
        bulk_insert_json['name'] = bulk_title
        bulk_insert_json['status'] =  Bulk_Status.IN_QUEUE.value
        bulk_insert_json['request_type'] = request_type
        bulk_insert_json['total_count'] = len(business_names_list)
        bulk_insert_json['passed_count'] = 0
        bulk_insert_json['user'] = str(request.user.id)
        

        bulk_id = None
        while True:
            try:
                collection_Bulks.insert_one(bulk_insert_json)

                bulk_id = collection_Bulks.find_one(bulk_insert_json)['_id']
                print(bulk_id)
                if (bulk_id != None):
                    break
            except Exception as e:
                print(e)

                continue

        # process_num = 10
        # chunks = chunkIt(business_names_list, process_num)
        
        
        # processes = []


        # for i in range(len(chunks)):
        #     processes.append(multiprocessing.Process(target=insert_bulk_requests_list_multiprocess_mongodb, args=[request, bulk_id, request_type, chunks[i]]))

        # for p in processes:
        #     p.start()

        # for p in processes:
        #     p.join()        

        insert_bulk_requests_list_multiprocess_mongodb(request, bulk_id, request_type, business_names_list)
        # update remain in User_Other_Fields in db
        update_remain_count_user(request, -1 * len(business_names_list))

    except Exception as e:
        print(e)
        pass

def get_API_response(input_text, request_type):

    try:

        if request_type == 'Youtube-Category Extraction':

            url = "http://" + str(IP_SINGLE_API_CATEGORY) + ":8943/CA_single?request_type=" + request_type +"&text=" + input_text

            payload = ""
            headers = {
                'content-type': "application/json"
            }

            response = requests.request("GET", url, data=payload, headers=headers)
            if ('result' in response.json()):
                return response.json()['result'], Request_Status.SUCCESS
        else:
            url = "http://" + str(IP_SINGLE_API) + ":8942/CA_single?request_type=" + request_type +"&text=" + input_text

            payload = ""
            headers = {
                'content-type': "application/json"
            }

            response = requests.request("GET", url, data=payload, headers=headers)
            if ('result' in response.json()):
                return response.json()['result'], Request_Status.SUCCESS
                
    except Exception as e:
        print(e)
        return 'not found', Request_Status.FAILED

def get_google_spreadsheet_response(column_name, google_spreadsheet_url):

    try:
        
        url = "http://138.201.111.134:8887/google-sheet?column_name=" + column_name + "&google_spreadsheets_url=" + google_spreadsheet_url
        payload = ""
        headers = {
            'content-type': "application/json"
        }

        response = requests.request("GET", url, data=payload, headers=headers)
        return response.text , Request_Status.SUCCESS
    except Exception as e:
        print(e)
        return '', Request_Status.FAILED

def insert_plan_db (title, plan_status, description, count, price, discount, order, tag):
    
    try:
        modify_time = datetime.datetime.now()

        plan_obj = Plan( modify_time_slot = modify_time, name=title, status=plan_status, description = description, count = count, price=price, discount = discount, order= order, tag =tag)
        plan_obj.save()

        print(colored('Plan insert successfully from client admin !!!', 'blue'))
        return 0
    except Exception as e:
        print(e)
        return 0

def insert_user_log_db (user, tag):
    
    try:
        # log_obj = User_Log( time_slot = datetime.datetime.now(), activity=tag, user= user)
        # log_obj.save()

        collection_User_Log.insert_one({'time_slot': datetime.datetime.now(), 'activity': tag, 'user': str(user.id)})


        print(colored('User_Log insert successfully from client !!!', 'blue'))
        return 0
    except Exception as e:
        print(e)
        return 0

def insert_payment_db (request, status, p_type, price, plan, paying_id):
    
    try:
        payment_obj = Payment( payment_time_slot = datetime.datetime.now(), status=status, payment_type=p_type, price = price, plan = plan, paying_id = paying_id, user = request.user)
        payment_obj.save()

        print(colored('Payment insert successfully from client user !!!', 'blue'))
        return 0
    except Exception as e:
        print(e)
        return 0

def insert_ticket_db_client (user, title):
    
    try:
        ticket_obj = Ticket(answer_time_slot= datetime.datetime.now(), question_time_slot= datetime.datetime.now(),client_seen_status= Seen_Status.SEEN.value, admin_seen_status = Seen_Status.UNSEEN.value, title=title, empty_status=Empty_Status.EMPTY.value, status=Ticket_Status.USER_CREATED.value, user = user)
        ticket_obj.save()

        print(colored('Ticket insert successfully from client user !!!', 'blue'))
        return ticket_obj.id
    except Exception as e:
        print(str(e))
        return 0

def insert_ticket_db_admin (user, title):
    
    try:
        ticket_obj = Ticket(answer_time_slot= datetime.datetime.now(), question_time_slot= datetime.datetime.now(),client_seen_status= Seen_Status.UNSEEN.value, admin_seen_status = Seen_Status.SEEN.value, title=title, empty_status=Empty_Status.NOT_EMPTY.value, status=Ticket_Status.ADMIN_NOTIFICATION.value, user = user)
        ticket_obj.save()

        print(colored('Ticket insert successfully from client user !!!', 'blue'))
        return ticket_obj.id
    except Exception as e:
        print(str(e))
        return 0

def insert_message_db (user, body, ticket, user_type):
    
    try:
        message_obj = Ticket_Message(ticket = ticket, time_slot= datetime.datetime.now(), user_type= user_type, body=body, seen_status=Seen_Status.UNSEEN.value, user = user)
        message_obj.save()

        print(colored('Message insert successfully from client user !!!', 'blue'))
        return 0
    except Exception as e:
        print(str(e))
        return 0

def update_remain_count_user(request, value):
    try:
        user_other_fields_obj = User_Other_Fields.objects.filter(user_id=request.user.id)[0]
        user_other_fields_obj.remain_count = user_other_fields_obj.remain_count + value
        user_other_fields_obj.save()
    except Exception as e:
        print(colored(str(e), 'red'))

def update_plan_user(request, plan_obj, promotion_code_used):
    try:
        user_other_fields_obj = User_Other_Fields.objects.filter(user_id=request.user.id)[0]
        user_other_fields_obj.plan = plan_obj
        if (promotion_code_used):
            user_other_fields_obj.promotion_code_used = promotion_code_used
        user_other_fields_obj.save()
    except Exception as e:
        print(colored(str(e), 'red'))

def check_remain_count_user(request, value):
    try:
        if (request.user.is_superuser): # ignore check remain for superuser
            return True, 1000000

        user_other_fields_obj = User_Other_Fields.objects.filter(user_id=request.user.id)[0]
        if (user_other_fields_obj.remain_count >= value):
            return True, user_other_fields_obj.remain_count
        else:
            return False, user_other_fields_obj.remain_count

    except Exception as e:
        print(colored(str(e), 'red'))
        return False, 0

def insert_single_request_db (request, business_name, request_type):
    
    try:
        request_time = datetime.datetime.now()
        result, status = get_API_response(business_name, request_type)
        result_time = datetime.datetime.now()

        request_obj = Request( request_time_slot = request_time, result_time_slot=result_time, query = business_name, result=result, status = status.value, request_type = request_type, run_type= Request_Run_Type.SINGLE.value, bulk=None, user= request.user)
        request_obj.save()

        # update remain in User_Other_Fields in db
        update_remain_count_user(request, -1)

        print(colored('Request Single insert successfully from client user !!!', 'blue'))

        return result

    except Exception as e:
        print(e)
        return ''

def insert_single_request_db_mongodb (request, text_query, request_type):
    
    while True:
        try:
            request_time = datetime.datetime.now()
            result, status = get_API_response(text_query, request_type)
            result_time = datetime.datetime.now()

            request_insert_json= {}
            request_insert_json['request_time_slot'] = request_time
            request_insert_json['result_time_slot'] = result_time
            request_insert_json['query'] = text_query
            request_insert_json['result'] = result
            request_insert_json['status'] = status.value
            request_insert_json['request_type'] = request_type
            request_insert_json['run_type'] = Request_Run_Type.SINGLE.value
            request_insert_json['bulk'] = None
            request_insert_json['user'] =  str(request.user.id)

            collection_Requests.insert_one(request_insert_json)
            
            request_id = collection_Requests.find_one(request_insert_json)['_id']
            if (request_id == None):
                continue
            # update remain in User_Other_Fields in db
            update_remain_count_user(request, -1)

            print(colored('Request Single insert successfully from client user !!!', 'blue'))

            return result

        except Exception as e:
            print(e)
            continue

def insert_demo_request_db_mongodb (request, business_name, request_type):
    
    while True:
        try:
            request_time = datetime.datetime.now()
            result, status = get_API_response(business_name, request_type)
            result_time = datetime.datetime.now()

            request_insert_json= {}
            request_insert_json['request_time_slot'] = request_time
            request_insert_json['result_time_slot'] = result_time
            request_insert_json['query'] = business_name
            request_insert_json['result'] = result
            request_insert_json['status'] = status.value
            request_insert_json['request_type'] = request_type
            request_insert_json['run_type'] = Request_Run_Type.DEMO.value
            request_insert_json['bulk'] = None
            request_insert_json['user'] =  None

            collection_Requests.insert_one(request_insert_json)
            
            request_id = collection_Requests.find_one(request_insert_json)['_id']
            if (request_id == None):
                continue
            
            print(colored('Request Demo insert successfully from Homepage !!!', 'blue'))

            return result

        except Exception as e:
            print(e)
            continue

def insert_demo_request_db (request, business_name, request_type):
    
    while True:
        try:
            request_time = datetime.datetime.now()
            result, status = get_API_response(business_name, request_type)
            result_time = datetime.datetime.now()



            request_obj = Request( request_time_slot = request_time, result_time_slot=result_time, query = business_name, result=result, status = status.value, request_type = request_type, run_type= Request_Run_Type.DEMO.value, bulk=None, user= None)
            request_obj.save()


            print(colored('Request DEMO insert successfully from client user !!!', 'blue'))

            return result

        except Exception as e:
            print(e)
            continue

def clean_data_input(business_names_list):
    out_list = []
    for item in business_names_list:
        try:
            if (item.strip() != ''):
                out_list.append(item.strip())
        except Exception as e:
            print(e)
            continue
    return out_list

def get_users_statistics():

    total_users = 0
    monthly_avg_users = 0

    monthly_users_list_value = [0] * MONTH_COUNT_REPORT # 12 months of the year
    monthly_users_list_month = ["Jan"] * MONTH_COUNT_REPORT # 12 months of the year

    monthly_users_list_json = { "data" :monthly_users_list_value, "months": monthly_users_list_month}

    row_list_all = User.objects.all()

    now = datetime.datetime.now()

    # get list of last 12 months
    month_list_recent = [now.strftime("%B")[:3]]
    for _ in range(1, MONTH_COUNT_REPORT):
        now = now.replace(day=1) - datetime.timedelta(days=1)
        month_list_recent.append(str(now.strftime("%B")[:3]))
    
    monthly_users_list_json['months'] = list(reversed(month_list_recent))

    for index, row in enumerate(list(row_list_all)):
        try:
            total_users += 1

            diff_month_value = diff_month(datetime.datetime.now(), row.date_joined)
            if (diff_month_value < MONTH_COUNT_REPORT):
                monthly_users_list_json['data'][MONTH_COUNT_REPORT-diff_month_value-1] += 1

        except Exception as e:
            print(e)
            continue


    count_non_zero_users_month = 0
    for month_users in monthly_users_list_json['data']:
        if (month_users != 0):
            count_non_zero_users_month += 1

    if (count_non_zero_users_month != 0):
        monthly_avg_users = total_users / count_non_zero_users_month
    else:
        monthly_avg_users =0

    return total_users, monthly_avg_users, monthly_users_list_json

def insert_post_db_admin (title, image_name):
    
    try:
        post_obj = Post(time_slot= datetime.datetime.now(), title = title, image_name = image_name, status= Post_Status.UNPUBLISHED.value )
        post_obj.save()

        print(colored('Post insert successfully from admin user !!!', 'blue'))
        return post_obj.id
    except Exception as e:
        print(str(e))
        return 0

def insert_discount_db_admin (code, value):
    
    try:
        discount_obj = Discount_Code(time_slot= datetime.datetime.now(), code = code, value = int(value), status= Plan_Status.DISABLE.value )
        discount_obj.save()

        print(colored('Discount insert successfully from admin user !!!', 'blue'))
        return discount_obj.id
    except Exception as e:
        print(str(e))
        return 0

def insert_paragraph_db (body, post):
    
    try:
        paragraph_obj = Post_Paragraph(post=post, time_slot= datetime.datetime.now(), body=body)
        paragraph_obj.save()

        print(colored('Paragraph insert successfully from Admin user !!!', 'blue'))
        return 0
    except Exception as e:
        print(str(e))
        return 0

def check_discount_code(discount_code):

    discount_list_all = Discount_Code.objects.filter(status=Plan_Status.ENABLE.value)
    for discount in discount_list_all:
        if discount.code == discount_code:
            return discount.value
    return 0

############################################# Client ##########################################################
@login_required(login_url="/login/")
def dashboard_client_weekly(request):

    insert_user_log_db (request.user, 'Dashboard')
    valid_to_submit, remain_count = check_remain_count_user(request, 1)
    row_list_request_single, total_count_request_single = get_request_list_page_mongodb(request, 0, ROW_LIST_SHOW_COUNT, request.user, Request_Run_Type.SINGLE)
    row_list_request_bulk, total_count_request_bulk = get_request_list_page_mongodb(request, 0, ROW_LIST_SHOW_COUNT, request.user, Request_Run_Type.BULK)
    
    row_list_ticket, total_count_ticket, unread_count_ticket = get_ticket_list_page(request, 0, ROW_LIST_SHOW_COUNT, request.user)

    plan_name = get_user_plan_name(request.user.id)

    monthly_request = get_week_list_request (request, None)

    return render(request, "home/dashboard_client.html", {"msg": 'SUCCESS',
                                                          "remain_count" : remain_count,
                                                          "total_count_request_single" : total_count_request_single,
                                                          "total_count_request_bulk" : total_count_request_bulk,
                                                          "total_count_ticket" : total_count_ticket,
                                                          "unread_count_ticket" : unread_count_ticket,
                                                          "monthly_request" : monthly_request,
                                                          "report_period" : "week",
                                                          "plan": plan_name
                                                        })

@login_required(login_url="/login/")
def ner_description_doc(request):
     return render(request, "home/ner_description_doc.html", {"msg": 'SUCCESS',
                                                          "segment": 'ner-doc'
                                                        })

@login_required(login_url="/login/")
def category_description_doc(request):
     return render(request, "home/category_description_doc.html", {"msg": 'SUCCESS',
                                                          "segment": 'category-doc'
                                                        })

@login_required(login_url="/login/")
def dashboard_client(request):

    insert_user_log_db (request.user, 'Dashboard')
    valid_to_submit, remain_count = check_remain_count_user(request, 1)
    row_list_request_single, total_count_request_single = get_request_list_page_mongodb(request, 0, ROW_LIST_SHOW_COUNT, request.user, Request_Run_Type.SINGLE)
    row_list_request_bulk, total_count_request_bulk = get_request_list_page_mongodb(request, 0, ROW_LIST_SHOW_COUNT, request.user, Request_Run_Type.BULK)

    row_list_ticket, total_count_ticket, unread_count_ticket = get_ticket_list_page(request, 0, ROW_LIST_SHOW_COUNT, request.user)

    plan_name = get_user_plan_name(request.user.id)

    monthly_request =  get_month_list_request(request, None)

    return render(request, "home/dashboard_client.html", {"msg": 'SUCCESS',
                                                          "remain_count" : remain_count,
                                                          "total_count_request_single" : total_count_request_single,
                                                          "total_count_request_bulk" : total_count_request_bulk,
                                                          "total_count_ticket" : total_count_ticket,
                                                          "unread_count_ticket" : unread_count_ticket,
                                                          "monthly_request" : monthly_request,
                                                          "report_period" : "month",
                                                          "plan": plan_name
                                                        })

@login_required(login_url="/login/")
def profile_client(request):

    msg = ''
    insert_user_log_db (request.user, 'Profile')
    valid_to_submit, remain_count = check_remain_count_user(request, 1)
    row_list_request_single, total_count_request_single = get_request_list_page_mongodb(request, 0, ROW_LIST_SHOW_COUNT, request.user, Request_Run_Type.SINGLE)
    row_list_request_bulk, total_count_request_bulk = get_request_list_page_mongodb(request, 0, ROW_LIST_SHOW_COUNT, request.user, Request_Run_Type.BULK)

    row_list_ticket, total_count_ticket, unread_count_ticket = get_ticket_list_page(request, 0, ROW_LIST_SHOW_COUNT, request.user)

    plan_name = get_user_plan_name(request.user.id)

    user_key= ''
    try:
        user_other_fields_obj = User_Other_Fields.objects.filter(user_id=request.user.id)[0]
        user_key = user_other_fields_obj.user_key
    except Exception as e:
        print(colored(str(e), 'red'))
    

    return render(request, "home/profile_client.html", {"msg": msg,
                                                          "remain_count" : remain_count,
                                                          "total_count_request_single" : total_count_request_single,
                                                          "total_count_request_bulk" : total_count_request_bulk,
                                                          "total_count_ticket" : total_count_ticket,
                                                          "unread_count_ticket" : unread_count_ticket,
                                                          "email" : request.user.email,
                                                          "user_key" : user_key,
                                                          "username" : request.user.username,
                                                          "plan": plan_name
                                                        })

@login_required(login_url="/login/")
def change_password_client(request):
    msg = None
    success = False


    valid_to_submit, remain_count = check_remain_count_user(request, 1)
    row_list_request_single, total_count_request_single = get_request_list_page_mongodb(request, 0, ROW_LIST_SHOW_COUNT, request.user, Request_Run_Type.SINGLE)
    row_list_request_bulk, total_count_request_bulk = get_request_list_page_mongodb(request, 0, ROW_LIST_SHOW_COUNT, request.user, Request_Run_Type.BULK)

    row_list_ticket, total_count_ticket, unread_count_ticket = get_ticket_list_page(request, 0, ROW_LIST_SHOW_COUNT, request.user)

    plan_name = get_user_plan_name(request.user.id)

    user_key= ''
    try:
        user_other_fields_obj = User_Other_Fields.objects.filter(user_id=request.user.id)[0]
        user_key = user_other_fields_obj.user_key
    except Exception as e:
        print(colored(str(e), 'red'))
        
    if request.method == "POST": # just in confirm new bulk data buttton clicked

        ############################################ Changing Password #########################################
        try:
            # current_password = request.POST.get('current_password').strip()
            email = request.POST.get('email').strip()
        

            new_password_1 = request.POST.get('new_password_1').strip()
            new_password_2 = request.POST.get('new_password_2').strip()

            if (new_password_1 != new_password_2):
                msg = 'new passwords are not the same !'
                return render(request, "home/profile_client.html", {"msg": msg,
                                                                    "remain_count" : remain_count,
                                                                    "total_count_request_single" : total_count_request_single,
                                                                    "total_count_request_bulk" : total_count_request_bulk,
                                                                    "total_count_ticket" : total_count_ticket,
                                                                    "unread_count_ticket" : unread_count_ticket,
                                                                    "email" : request.user.email,
                                                                    "user_key" : user_key,
                                                                    "username" : request.user.username,
                                                                    "plan": plan_name
                                                                    })

            try:
                password_validation.validate_password(new_password_1)
            except Exception as e:
                msg = str(e)
                return render(request, "home/profile_client.html", {"msg": msg,
                                                                    "remain_count" : remain_count,
                                                                    "total_count_request_single" : total_count_request_single,
                                                                    "total_count_request_bulk" : total_count_request_bulk,
                                                                    "total_count_ticket" : total_count_ticket,
                                                                    "unread_count_ticket" : unread_count_ticket,
                                                                    "email" : request.user.email,
                                                                    "user_key" : user_key,
                                                                    "username" : request.user.username,
                                                                    "plan": plan_name
                                                                    })

            request.user.set_password(new_password_1)
            request.user.save()

            msg = 'user password changed successfully ! '
            success = True

            return render(request, "home/profile_client.html", {"msg": msg,
                                                        "remain_count" : remain_count,
                                                        "total_count_request_single" : total_count_request_single,
                                                        "total_count_request_bulk" : total_count_request_bulk,
                                                        "total_count_ticket" : total_count_ticket,
                                                        "unread_count_ticket" : unread_count_ticket,
                                                        "email" : request.user.email,
                                                        "user_key" : user_key,
                                                        "username" : request.user.username,
                                                        "plan": plan_name
                                                    })


        except Exception as e:
            print(e)
            msg= str(e)
            success = False

            
    return render(request, "home/profile_client.html", {"msg": msg, "success": success,
                                                        "remain_count" : remain_count,
                                                        "total_count_request_single" : total_count_request_single,
                                                        "total_count_request_bulk" : total_count_request_bulk,
                                                        "total_count_ticket" : total_count_ticket,
                                                        "unread_count_ticket" : unread_count_ticket,
                                                        "email" : request.user.email,
                                                        "user_key" : user_key,
                                                        "username" : request.user.username,
                                                        "plan": plan_name
                                                        })

@login_required(login_url="/login/")
def requests_new_bulk(request):

    insert_user_log_db (request.user, 'New_Bulk')

    requests_types_tuple = list(Request_Type.choices())
    requests_types = list((i[1]) for i in requests_types_tuple) 

    return render(request, "home/requests-new_bulk.html", {"msg": 'SUCCESS' , "segment": 'requests-new', "requests_types": requests_types})

@login_required(login_url="/login/")
def requests_sentiment_single_client(request):

    valid_to_submit = True
    remain_count = 0

    insert_user_log_db (request.user, 'Single_Sentiment_Requests')

    current_pagination = 1    
    try:
        current_pagination = int (request.path.split('/P')[-1])
    except Exception as e:
        pass
    
    requests_types_tuple = list(Request_Type.choices())
    requests_types = list((i[1]) for i in requests_types_tuple)
    request_result = ''
    try:
        if request.method == "POST": 
            text_query = request.POST.get('input_text_query').strip()
            request_type = request.POST.get('input_request_type').strip()

            # insert single request mongo
            valid_to_submit, remain_count = check_remain_count_user(request, 1)
            if (valid_to_submit):
                request_result = insert_single_request_db_mongodb(request, text_query, request_type)
        else:
            valid_to_submit, remain_count = check_remain_count_user(request, 1)

 
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_request, total_count_request = get_request_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, Request_Run_Type.SINGLE, Request_Type.SENTIMENT_ANALYSIS.value)

        last_pagination = 1
        if (total_count_request == 0):
            last_pagination = 1
        elif (total_count_request % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/requests-sentiment-single_client.html", {"msg": 'SUCCESS',
                                                                "valid_to_submit" : valid_to_submit,
                                                                "remain_count" : remain_count,
                                                                "segment": 'sentiment-single', 
                                                                "requests_types": requests_types,
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_request),
                                                                "total_items": total_count_request,
                                                                "request_list": row_list_request,
                                                                "request_result": request_result})

@login_required(login_url="/login/")
def requests_ner_single_client(request):

    valid_to_submit = True
    remain_count = 0

    insert_user_log_db (request.user, 'Single_NER_Requests')

    current_pagination = 1    
    try:
        current_pagination = int (request.path.split('/P')[-1])
    except Exception as e:
        pass
    
    requests_types_tuple = list(Request_Type.choices())
    requests_types = list((i[1]) for i in requests_types_tuple)
    request_result = ''
    try:
        if request.method == "POST": 
            text_query = request.POST.get('input_text_query').strip()
            request_type = request.POST.get('input_request_type').strip()

            # insert single request mongo
            valid_to_submit, remain_count = check_remain_count_user(request, 1)
            if (valid_to_submit):
                request_result = insert_single_request_db_mongodb(request, text_query, request_type)
        else:
            valid_to_submit, remain_count = check_remain_count_user(request, 1)

 
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_request, total_count_request = get_request_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, Request_Run_Type.SINGLE, Request_Type.NAMED_ENTITY_RECOGNITION.value)

        last_pagination = 1
        if (total_count_request == 0):
            last_pagination = 1
        elif (total_count_request % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/requests-ner-single_client.html", {"msg": 'SUCCESS',
                                                                "valid_to_submit" : valid_to_submit,
                                                                "remain_count" : remain_count,
                                                                "segment": 'ner-single', 
                                                                "requests_types": requests_types,
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_request),
                                                                "total_items": total_count_request,
                                                                "request_list": row_list_request,
                                                                "request_result": request_result})

@login_required(login_url="/login/")
def requests_keyword_single_client(request):

    valid_to_submit = True
    remain_count = 0

    insert_user_log_db (request.user, 'Single_Keyword_Requests')

    current_pagination = 1    
    try:
        current_pagination = int (request.path.split('/P')[-1])
    except Exception as e:
        pass
    
    requests_types_tuple = list(Request_Type.choices())
    requests_types = list((i[1]) for i in requests_types_tuple)
    request_result = ''
    try:
        if request.method == "POST": 
            text_query = request.POST.get('input_text_query').strip()
            request_type = request.POST.get('input_request_type').strip()

            # insert single request mongo
            valid_to_submit, remain_count = check_remain_count_user(request, 1)
            if (valid_to_submit):
                request_result = insert_single_request_db_mongodb(request, text_query, request_type)
        else:
            valid_to_submit, remain_count = check_remain_count_user(request, 1)

 
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_request, total_count_request = get_request_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, Request_Run_Type.SINGLE, Request_Type.KEYWORD_EXTRACTION.value)

        last_pagination = 1
        if (total_count_request == 0):
            last_pagination = 1
        elif (total_count_request % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/requests-keyword-single_client.html", {"msg": 'SUCCESS',
                                                                "valid_to_submit" : valid_to_submit,
                                                                "remain_count" : remain_count,
                                                                "segment": 'keyword-single', 
                                                                "requests_types": requests_types,
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_request),
                                                                "total_items": total_count_request,
                                                                "request_list": row_list_request,
                                                                "request_result": request_result})

@login_required(login_url="/login/")
def requests_category_single_client(request):

    valid_to_submit = True
    remain_count = 0

    insert_user_log_db (request.user, 'Single_Category_Requests')

    current_pagination = 1    
    try:
        current_pagination = int (request.path.split('/P')[-1])
    except Exception as e:
        pass
    
    requests_types_tuple = list(Request_Type.choices())
    requests_types = list((i[1]) for i in requests_types_tuple)
    request_result = ''
    try:
        if request.method == "POST": 
            text_query = request.POST.get('input_text_query').strip()
            request_type = request.POST.get('input_request_type').strip()
            # insert single request mongo
            valid_to_submit, remain_count = check_remain_count_user(request, 1)
            if (valid_to_submit):
                request_result = insert_single_request_db_mongodb(request, text_query, request_type)
        else:
            valid_to_submit, remain_count = check_remain_count_user(request, 1)

 
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_request, total_count_request = get_request_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, Request_Run_Type.SINGLE, Request_Type.YOUTUBE_CATEGORY_EXTRACTION.value)

        last_pagination = 1
        if (total_count_request == 0):
            last_pagination = 1
        elif (total_count_request % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/requests-category-single_client.html", {"msg": 'SUCCESS',
                                                                "valid_to_submit" : valid_to_submit,
                                                                "remain_count" : remain_count,
                                                                "segment": 'category-single', 
                                                                "requests_types": requests_types,
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_request),
                                                                "total_items": total_count_request,
                                                                "request_list": row_list_request,
                                                                "request_result": request_result})

@login_required(login_url="/login/")
def requests_bulk_client(request):

    try:
        insert_user_log_db (request.user, 'Bulk_Requests')

        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_request, total_count_request = get_request_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, Request_Run_Type.BULK)

        last_pagination = 1
        if (total_count_request == 0):
            last_pagination = 1
        elif (total_count_request % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/requests-bulk_client.html", {"msg": 'SUCCESS',
                                                                "segment": 'requests-bulk_client', 
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_request),
                                                                "total_items": total_count_request,
                                                                "request_list": row_list_request})

@login_required(login_url="/login/")
def requests_new_bulk_data_confirm(request):

    remain_count = 0
    business_names_list = []
    valid_to_submit = True

    try:
        insert_user_log_db (request.user, 'Bulk_Data_Confirm')
        print(colored(request.method, 'blue'))
        if request.method == "POST": 
            bulk_title = request.POST.get('input_bulk_title').strip()
            request_type = request.POST.get('input_request_type').strip()

            business_names_raw = ''

            input_method = request.path.split('/method_')[-1]
            if ('json' in input_method):
                try:
                    json_file_name = None
                    try:
                        json_file_name = request.FILES["json_file"]
                        json_file = json_file_name.read()
                        for comment in json.loads(json_file):
                            try:
                                business_names_list.append(comment['Msg'])
                            except Exception as e:
                                print(e)
                        
                        for item in business_names_list:
                            if (business_names_raw == ''):
                                business_names_raw += item.replace(SPLIT_TOKEN, '').strip()
                            else:
                                business_names_raw += SPLIT_TOKEN + item.replace(SPLIT_TOKEN, '').strip()

                    except Exception as e:
                        print(e)
                        pass

                    
                    if (json_file_name == None):
                        business_names_raw = request.POST.get('input_business_names').strip()
                        business_names_list = business_names_raw.split(SPLIT_TOKEN)
                except Exception as e:
                    print(e)

            elif ('google_spreadsheet' in input_method):
                try:
                    
                    if (request.POST.get('url_spreadsheet') == None and request.POST.get('column_name_spreadsheet')==None):
                        business_names_raw = request.POST.get('input_business_names').strip()
                        business_names_list = business_names_raw.split(SPLIT_TOKEN)
                    else:
                        url_spreadsheet = request.POST.get('url_spreadsheet').strip() 
                        column_name = request.POST.get('column_name_spreadsheet').strip()
                        response_text , response_status = get_google_spreadsheet_response(column_name, url_spreadsheet)

                        business_names_list = list(response_text[5:-4].split('",\n  "'))
                        if (len ( business_names_list ) == 1):
                            business_names_list = list(business_names_list[0].split('", "'))

                        business_names_list = clean_data_input(business_names_list)
                        for item in business_names_list:
                            if (business_names_raw == ''):
                                business_names_raw += item.replace(SPLIT_TOKEN, '').strip()
                            else:
                                business_names_raw += SPLIT_TOKEN + item.replace(SPLIT_TOKEN, '').strip()

                except Exception as e:
                    print(e)
            
            elif ('excel' in input_method):
                try:
                    excel_file_name_xlsx = None
                    try:
                        excel_file_name_xlsx = request.FILES["excel_file_xlsx"]
                    except Exception as e:
                        pass
                    excel_file_name_csv = None
                    try:
                        excel_file_name_csv = request.FILES["excel_file_csv"]
                    except Exception as e:
                        pass

                    excel_file_name = None
                    try:
                        excel_file_name = request.FILES["excel_file_xlsx"]
                    except Exception as e:
                        pass

                    if (excel_file_name_xlsx == None and excel_file_name_csv == None):
                        business_names_raw = request.POST.get('input_business_names').strip()
                        business_names_list = business_names_raw.split(SPLIT_TOKEN)

                    # excel .xlsx
                    elif (excel_file_name and str(excel_file_name).endswith('.xlsx')):

                        column_name = request.POST.get('column_name_excel_xlsx').strip()
                        # you may put validations here to check extension or file size
                        wb = openpyxl.load_workbook(excel_file_name)

                        # getting a particular sheet by name out of many sheets
                        worksheet = wb["Sheet1"]

                        # iterating over the rows and
                        # getting value from each cell in row
                        row_index = 0
                        selected_column = -1

                        for row in worksheet.iter_rows():
                            if (row_index == 0):
                                row_index += 1
                                column = 0
                                for item in row:
                                    if (item.value == column_name):
                                        selected_column = column
                                    column += 1 
                                continue
                            
                            # column_name input not found
                            if selected_column == -1:
                                break

                            if (selected_column < len(row)):
                                business_names_list.append(row[selected_column].value)

                            row_index += 1

                        business_names_list = clean_data_input(business_names_list)
                        business_names_raw = ''
                        for item in business_names_list:
                            if (business_names_raw == ''):
                                business_names_raw += item.replace(SPLIT_TOKEN, '').strip()
                            else:
                                business_names_raw += SPLIT_TOKEN + item.replace(SPLIT_TOKEN, '').strip()

                       


                    # excel .csv
                    else:
                        
                        try:
                            excel_file_name = request.FILES["excel_file_csv"]
                        except Exception as e:
                            pass

                        if (excel_file_name and str(excel_file_name).endswith('.csv')):

                            column_name = request.POST.get('column_name_excel_csv').strip()
                            delimiter = request.POST.get('delimiter_excel_csv').strip()

                            data = pd.read_csv(excel_file_name, header=None, sep=delimiter)
                            
                            selected_index = -1

                            for index in range(len(data.columns)):
                                if (data[index][0] == column_name):
                                    selected_index = index

                            if (selected_index >=0):
                                for ind, row in enumerate(data[selected_index]):
                                    if (ind == 0):
                                        continue
                                    business_names_list.append(row)
                            print(business_names_list)
                            business_names_list = clean_data_input(business_names_list)
                            business_names_raw = ''
                            for item in business_names_list:
                                if (business_names_raw == ''):
                                    business_names_raw += item.replace(SPLIT_TOKEN, '').strip()
                                else:
                                    business_names_raw += SPLIT_TOKEN + item.replace(SPLIT_TOKEN, '').strip()
                                
 
                            pass

                except Exception as e:
                    print(e)


            # check total count length
            if (len(business_names_list) > BULK_SIZE_LIMIT):
                business_names_list = business_names_list[0:BULK_SIZE_LIMIT]
                business_names_raw = ''
                for item in business_names_list:
                    business_names_raw += SPLIT_TOKEN + item.replace(SPLIT_TOKEN,' ').strip()

            valid_to_submit, remain_count = check_remain_count_user(request, len(business_names_list))
            
        else:

            business_names_raw = ''
            bulk_title = ''
            request_type = ''
            business_names_list = []



        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_bulk, total_count_names = get_business_names_list_page(business_names_list, start, ROW_LIST_SHOW_COUNT)

        


        last_pagination = 1
        if (total_count_names == 0):
            last_pagination = 1
        elif (total_count_names % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_names / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_names / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/requests-new_bulk_data_confirm.html", {"msg": 'SUCCESS',
                                                                        "valid_to_submit" : valid_to_submit,
                                                                        "remain_count" : remain_count,
                                                                        "segment": 'requests-new',
                                                                        "bulk_title": bulk_title,
                                                                        "request_type": request_type,
                                                                        "current_pagination": current_pagination,
                                                                        "last_pagination": last_pagination,
                                                                        "page_limit": len(row_list_bulk),
                                                                        "total_items": total_count_names,
                                                                        "business_names_list_page": row_list_bulk,
                                                                        "business_names_raw": business_names_raw})

def chunkIt(seq, num):
    '''
    function_name: chunkIt
    input: list, number of parts
    output: list of lists
    description: split list with approximate num of elements in each parts
    '''
    avg = len(seq) / float(num)
    out = []
    last = 0.0

    eps = 0.00001
    while last < len(seq):
        
        out.append(seq[int(last):int(last + avg)])
        last += avg

    return out

@login_required(login_url="/login/")
def requests_bulk_status_client(request):

    try:
        insert_user_log_db (request.user, 'Bulk_Status')
        
        if request.method == "POST": # just in confirm new bulk data buttton clicked

            bulk_title = request.POST.get('input_bulk_title').strip()
            request_type = request.POST.get('input_request_type').strip()

            business_names_raw = request.POST.get('input_business_names').strip()
            business_names_list = business_names_raw.strip().split(SPLIT_TOKEN)
            business_names_list = clean_data_input(business_names_list)
            # insert bulk request sqlite
            # gevent.spawn(insert_bulk_request_db(request, bulk_title, request_type, business_names_list))
            # insert_bulk_request_db_mongodb(request, bulk_title, request_type, business_names_list)

            insert_bulk_request_db_mongodb(request, bulk_title, request_type, business_names_list)


        current_pagination = 1   
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass

        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT
        row_list_bulk, total_count_bulk = get_bulk_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user)

        last_pagination = 1
        if (total_count_bulk == 0):
            last_pagination = 1
        elif (total_count_bulk % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_bulk / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_bulk / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(colored(str(e), 'red'))
        pass

    return render(request, "home/requests-bulk_status_client.html", {"msg": 'SUCCESS',
                                                                        "segment": 'requests-bulk_status',
                                                                        "current_pagination": current_pagination,
                                                                        "last_pagination": last_pagination,
                                                                        "page_limit": len(row_list_bulk),
                                                                        "total_items": total_count_bulk,
                                                                        "bulk_list": row_list_bulk})

@login_required(login_url="/login/")
def requests_sentiment_analytics_client(request):

    try:
        insert_user_log_db (request.user, 'requests_sentiment_analytics')
        
        if request.method == "POST": # just in confirm new bulk data buttton clicked

            bulk_title = request.POST.get('input_bulk_title').strip()
            request_type = request.POST.get('input_request_type').strip()

            business_names_raw = request.POST.get('input_business_names').strip()
            business_names_list = business_names_raw.strip().split(SPLIT_TOKEN)
            business_names_list = clean_data_input(business_names_list)
            # insert bulk request sqlite
            # gevent.spawn(insert_bulk_request_db(request, bulk_title, request_type, business_names_list))
            # insert_bulk_request_db_mongodb(request, bulk_title, request_type, business_names_list)

            insert_bulk_request_db_mongodb(request, bulk_title, request_type, business_names_list)


        current_pagination = 1   
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass

        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT
        row_list_bulk, total_count_bulk = get_bulk_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, None, Request_Type.SENTIMENT_ANALYSIS)
        
        sentiment_bulk_list_chart = { 
            "data_positive" :[],
            "data_neutral" :[],
            "data_negative" :[], 
            "labels": []
        }

        # reversed to have recent values in right side of chart
        for bulk in reversed(row_list_bulk) :
            sentiment_bulk_list_chart["data_positive"].append(bulk["positive_percent"])
            sentiment_bulk_list_chart["data_neutral"].append(bulk["neutral_percent"])
            sentiment_bulk_list_chart["data_negative"].append(bulk["negative_percent"])
            sentiment_bulk_list_chart["labels"].append(bulk["title"])


        last_pagination = 1
        if (total_count_bulk == 0):
            last_pagination = 1
        elif (total_count_bulk % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_bulk / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_bulk / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(colored(str(e), 'red'))
        pass

    return render(request, "home/requests-sentiment_analytics_client.html", {"msg": 'SUCCESS',
                                                                        "segment": 'sentiment-analytics',
                                                                        "current_pagination": current_pagination,
                                                                        "last_pagination": last_pagination,
                                                                        "page_limit": len(row_list_bulk),
                                                                        "total_items": total_count_bulk,
                                                                        "sentiment_bulk_list_chart": sentiment_bulk_list_chart,
                                                                        "bulk_list": row_list_bulk})

@login_required(login_url="/login/")
def requests_ner_analytics_client(request):

    try:
        insert_user_log_db (request.user, 'requests_NER_analytics')
        
        if request.method == "POST": # just in confirm new bulk data buttton clicked

            bulk_title = request.POST.get('input_bulk_title').strip()
            request_type = request.POST.get('input_request_type').strip()

            business_names_raw = request.POST.get('input_business_names').strip()
            business_names_list = business_names_raw.strip().split(SPLIT_TOKEN)
            business_names_list = clean_data_input(business_names_list)
            # insert bulk request sqlite
            # gevent.spawn(insert_bulk_request_db(request, bulk_title, request_type, business_names_list))
            # insert_bulk_request_db_mongodb(request, bulk_title, request_type, business_names_list)

            insert_bulk_request_db_mongodb(request, bulk_title, request_type, business_names_list)


        current_pagination = 1   
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass

        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT
        row_list_bulk, total_count_bulk = get_bulk_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, None, Request_Type.NAMED_ENTITY_RECOGNITION)
        
        
        last_pagination = 1
        if (total_count_bulk == 0):
            last_pagination = 1
        elif (total_count_bulk % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_bulk / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_bulk / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(colored(str(e), 'red'))
        pass

    return render(request, "home/requests-ner_analytics_client.html", {"msg": 'SUCCESS',
                                                                        "segment": 'ner-analytics',
                                                                        "current_pagination": current_pagination,
                                                                        "last_pagination": last_pagination,
                                                                        "page_limit": len(row_list_bulk),
                                                                        "total_items": total_count_bulk,
                                                                        "bulk_list": row_list_bulk})

@login_required(login_url="/login/")
def requests_keyword_analytics_client(request):

    try:
        insert_user_log_db (request.user, 'requests_Keyword_analytics')
        
        if request.method == "POST": # just in confirm new bulk data buttton clicked

            bulk_title = request.POST.get('input_bulk_title').strip()
            request_type = request.POST.get('input_request_type').strip()

            business_names_raw = request.POST.get('input_business_names').strip()
            business_names_list = business_names_raw.strip().split(SPLIT_TOKEN)
            business_names_list = clean_data_input(business_names_list)
            # insert bulk request sqlite
            # gevent.spawn(insert_bulk_request_db(request, bulk_title, request_type, business_names_list))
            # insert_bulk_request_db_mongodb(request, bulk_title, request_type, business_names_list)

            insert_bulk_request_db_mongodb(request, bulk_title, request_type, business_names_list)


        current_pagination = 1   
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass

        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT
        row_list_bulk, total_count_bulk = get_bulk_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, None, Request_Type.KEYWORD_EXTRACTION)
        
        
        last_pagination = 1
        if (total_count_bulk == 0):
            last_pagination = 1
        elif (total_count_bulk % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_bulk / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_bulk / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(colored(str(e), 'red'))
        pass

    return render(request, "home/requests-keyword_analytics_client.html", {"msg": 'SUCCESS',
                                                                        "segment": 'keyword-analytics',
                                                                        "current_pagination": current_pagination,
                                                                        "last_pagination": last_pagination,
                                                                        "page_limit": len(row_list_bulk),
                                                                        "total_items": total_count_bulk,
                                                                        "bulk_list": row_list_bulk})

@login_required(login_url="/login/")
def requests_category_youtube_analytics_client(request):

    try:
        insert_user_log_db (request.user, 'requests_category_youtube_analytics')
        
        if request.method == "POST": # just in confirm new bulk data buttton clicked

            bulk_title = request.POST.get('input_bulk_title').strip()
            request_type = request.POST.get('input_request_type').strip()

            business_names_raw = request.POST.get('input_business_names').strip()
            business_names_list = business_names_raw.strip().split(SPLIT_TOKEN)
            business_names_list = clean_data_input(business_names_list)
            # insert bulk request sqlite
            # gevent.spawn(insert_bulk_request_db(request, bulk_title, request_type, business_names_list))
            # insert_bulk_request_db_mongodb(request, bulk_title, request_type, business_names_list)

            insert_bulk_request_db_mongodb(request, bulk_title, request_type, business_names_list)


        current_pagination = 1   
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass

        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT
        row_list_bulk, total_count_bulk = get_bulk_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, None, Request_Type.YOUTUBE_CATEGORY_EXTRACTION)
        
        
        last_pagination = 1
        if (total_count_bulk == 0):
            last_pagination = 1
        elif (total_count_bulk % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_bulk / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_bulk / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(colored(str(e), 'red'))
        pass

    return render(request, "home/requests-category_youtube_analytics_client.html", {"msg": 'SUCCESS',
                                                                        "segment": 'category-youtube-analytics',
                                                                        "current_pagination": current_pagination,
                                                                        "last_pagination": last_pagination,
                                                                        "page_limit": len(row_list_bulk),
                                                                        "total_items": total_count_bulk,
                                                                        "bulk_list": row_list_bulk})

@login_required(login_url="/login/")
def update_bulk_status_ajax_client(request):
    current_pagination = 1   
    try:
        current_pagination = int (request.path.split('/P')[-1])
    except Exception as e:
        pass
    start = (current_pagination-1) * ROW_LIST_SHOW_COUNT
    row_list_bulk, total_count_bulk = get_bulk_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user)

    if request.is_ajax and request.method == "GET":
        return render(request, 'includes/bulk_status_table_client.html', {'bulk_list':row_list_bulk})

@login_required(login_url="/login/")
def update_bulk_sentiment_analytics_ajax_client(request):
    current_pagination = 1   
    try:
        current_pagination = int (request.path.split('/P')[-1])
    except Exception as e:
        pass
    start = (current_pagination-1) * ROW_LIST_SHOW_COUNT
    row_list_bulk, total_count_bulk = get_bulk_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, None, Request_Type.SENTIMENT_ANALYSIS)

    if request.is_ajax and request.method == "GET":
        return render(request, 'includes/bulk_sentiment_analytics_table_client.html', {'bulk_list':row_list_bulk})

@login_required(login_url="/login/")
def requests_ner_analytics_bulk(request):
    bulk_id = 0
    bulk_id_string = '0'
    page_string = '1'

    try:
        bulk_id_string = request.path.split('/bulk_')[-1]
    except Exception as e:
        pass


    query_bulk_id = {'_id' : ObjectId(bulk_id_string)}
    records_bulks = collection_Bulks.find(query_bulk_id)
    bulk_id = bulk_id_string

    if (records_bulks.count() == 0):
        return

    bulk_selected = records_bulks[0]

    all_requests = get_request_list_bulk_mongodb(request, bulk_id)
    
    top_ten_requests = []
    if len(all_requests) > 10 :
        top_ten_requests = all_requests[0:10]
    else:
        top_ten_requests= all_requests

    entities_count = {}
    entities_positive_sentiment_count = {}
    entities_neutral_sentiment_count = {}
    entities_negative_sentiment_count = {}

    counter = 1
    for request_item in all_requests:
        try:
            counter = counter + 1
            sentiment_result = request_item['sentiment']
            for entity in list(request_item['result']):
                if (len(entity[0]) < 2):
                    continue

                entity_string = entity[0] + ' (' + entity[1] + ')'

                try:
                    if entity_string in entities_count:
                        entities_count[entity_string] = entities_count[entity_string] + 1 
                    else:
                        entities_count[entity_string] = 1
                except Exception as e:
                    print(e)
                    
                if (sentiment_result == 'Positive'):
                    try:
                        if entity_string in entities_positive_sentiment_count:
                            entities_positive_sentiment_count[entity_string] = entities_positive_sentiment_count[entity_string] + 1 
                        else:
                            entities_positive_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)
                
                elif (sentiment_result == 'Neutral'):
                    try:
                        if entity_string in entities_neutral_sentiment_count:
                            entities_neutral_sentiment_count[entity_string] = entities_neutral_sentiment_count[entity_string] + 1 
                        else:
                            entities_neutral_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)

                elif (sentiment_result == 'Negative'):
                    try:
                        if entity_string in entities_negative_sentiment_count:
                            entities_negative_sentiment_count[entity_string] = entities_negative_sentiment_count[entity_string] + 1 
                        else:
                            entities_negative_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)
                    
        except Exception as e:

            print(e)
            continue

    # print(colored(str(entities_count) , 'green'))
    # print(colored(str(entities_positive_sentiment_count) , 'green'))
    # print(colored(str(entities_neutral_sentiment_count) , 'green'))
    # print(colored(str(entities_negative_sentiment_count) , 'green'))

    sortedDict_count = sorted(entities_count.items(), key=lambda x:x[1], reverse=True)

    # if len(sortedDict_count) > 10:
    #     sortedDict_count = sortedDict_count[0:10]

    data_out_count = []
    data_out_label = []
    data_out_sentiment_positive = []
    data_out_sentiment_neutral = []
    data_out_sentiment_negative = []

    
    for sorted_item in sortedDict_count:
        data_out_count.append(entities_count[sorted_item[0]])
        data_out_label.append(sorted_item[0])
        
        if (len(data_out_label) == 10):
            break
        
        # Positive
        if (sorted_item[0] in entities_positive_sentiment_count):
            data_out_sentiment_positive.append(entities_positive_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_positive.append(0)
        
        # Neutral
        if (sorted_item[0] in entities_neutral_sentiment_count):
            data_out_sentiment_neutral.append(entities_neutral_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_neutral.append(0)
        
        # Negative
        if (sorted_item[0] in entities_negative_sentiment_count):
            data_out_sentiment_negative.append(entities_negative_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_negative.append(0)
        

    ner_bulk_list_chart_top_count = {
        "data_positive" :data_out_sentiment_positive,
        "data_neutral" :data_out_sentiment_neutral,
        "data_negative" :data_out_sentiment_negative, 
        "data_count" : data_out_count,
        "labels": data_out_label
    }

        
    return render(request, 'home/requests-ner_analytics_report.html', {"msg": 'SUCCESS',
                                                            "segment": 'ner-analytics',
                                                            "ner_bulk_list_chart_top_count" : ner_bulk_list_chart_top_count,
                                                            "bulk_id" : bulk_id,
                                                            "request_list": top_ten_requests})

@login_required(login_url="/login/")
def requests_category_youtube_analytics_bulk(request):
    bulk_id = 0
    bulk_id_string = '0'
    page_string = '1'

    try:
        bulk_id_string = request.path.split('/bulk_')[-1]
    except Exception as e:
        pass


    query_bulk_id = {'_id' : ObjectId(bulk_id_string)}
    records_bulks = collection_Bulks.find(query_bulk_id)
    bulk_id = bulk_id_string

    if (records_bulks.count() == 0):
        return

    bulk_selected = records_bulks[0]

    all_requests = get_request_list_bulk_mongodb(request, bulk_id)
    
    
    top_ten_requests= all_requests

    # top_ten_requests = []
    # if len(all_requests) > 10 :
    #     top_ten_requests = all_requests[0:10]
    # else:
    #     top_ten_requests= all_requests

    entities_count = {}
    entities_positive_sentiment_count = {}
    entities_neutral_sentiment_count = {}
    entities_negative_sentiment_count = {}

    counter = 1
    for request_item in all_requests:
        try:
            counter = counter + 1
            sentiment_result = request_item['sentiment']
            for entity in list(request_item['result']):
                
                entity_string = entity

                try:
                    if entity_string in entities_count:
                        entities_count[entity_string] = entities_count[entity_string] + 1 
                    else:
                        entities_count[entity_string] = 1
                except Exception as e:
                    print(e)
                    
                if (sentiment_result == 'Positive'):
                    try:
                        if entity_string in entities_positive_sentiment_count:
                            entities_positive_sentiment_count[entity_string] = entities_positive_sentiment_count[entity_string] + 1 
                        else:
                            entities_positive_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)
                
                elif (sentiment_result == 'Neutral'):
                    try:
                        if entity_string in entities_neutral_sentiment_count:
                            entities_neutral_sentiment_count[entity_string] = entities_neutral_sentiment_count[entity_string] + 1 
                        else:
                            entities_neutral_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)

                elif (sentiment_result == 'Negative'):
                    try:
                        if entity_string in entities_negative_sentiment_count:
                            entities_negative_sentiment_count[entity_string] = entities_negative_sentiment_count[entity_string] + 1 
                        else:
                            entities_negative_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)
                    
        except Exception as e:

            print(e)
            continue

    # print(colored(str(entities_count) , 'green'))
    # print(colored(str(entities_positive_sentiment_count) , 'green'))
    # print(colored(str(entities_neutral_sentiment_count) , 'green'))
    # print(colored(str(entities_negative_sentiment_count) , 'green'))

    sortedDict_count = sorted(entities_count.items(), key=lambda x:x[1], reverse=True)

    # if len(sortedDict_count) > 10:
    #     sortedDict_count = sortedDict_count[0:10]

    data_out_count = []
    data_out_label = []
    data_out_sentiment_positive = []
    data_out_sentiment_neutral = []
    data_out_sentiment_negative = []

    
    for sorted_item in sortedDict_count:
        if (len(sorted_item[0]) < 2):
            continue

        data_out_count.append(entities_count[sorted_item[0]])
        data_out_label.append(sorted_item[0])
        
        # if (len(data_out_label) == 10):
        #     break

        
        
        # Positive
        if (sorted_item[0] in entities_positive_sentiment_count):
            data_out_sentiment_positive.append(entities_positive_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_positive.append(0)
        
        # Neutral
        if (sorted_item[0] in entities_neutral_sentiment_count):
            data_out_sentiment_neutral.append(entities_neutral_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_neutral.append(0)
        
        # Negative
        if (sorted_item[0] in entities_negative_sentiment_count):
            data_out_sentiment_negative.append(entities_negative_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_negative.append(0)
        
    
    category_bulk_list_chart_top_count = {
        "data_positive" :data_out_sentiment_positive,
        "data_neutral" :data_out_sentiment_neutral,
        "data_negative" :data_out_sentiment_negative, 
        "data_count" : data_out_count,
        "labels": data_out_label
    }

    print(colored(category_bulk_list_chart_top_count, 'green'))

    return render(request, 'home/requests-category_youtube_analytics_report.html', {"msg": 'SUCCESS',
                                                            "segment": 'category-youtube-analytics',
                                                            "category_bulk_list_chart_top_count" : category_bulk_list_chart_top_count,
                                                            "bulk_id" : bulk_id,
                                                            "request_list": top_ten_requests})

@login_required(login_url="/login/")
def requests_keyword_analytics_bulk(request):
    bulk_id = 0
    bulk_id_string = '0'
    page_string = '1'

    try:
        bulk_id_string = request.path.split('/bulk_')[-1]
    except Exception as e:
        pass


    query_bulk_id = {'_id' : ObjectId(bulk_id_string)}
    records_bulks = collection_Bulks.find(query_bulk_id)
    bulk_id = bulk_id_string

    if (records_bulks.count() == 0):
        return

    bulk_selected = records_bulks[0]

    all_requests = get_request_list_bulk_mongodb(request, bulk_id)
    
    top_ten_requests = []
    if len(all_requests) > 10 :
        top_ten_requests = all_requests[0:10]
    else:
        top_ten_requests= all_requests

    keywords_count = {}
    keywords_positive_sentiment_count = {}
    keywords_neutral_sentiment_count = {}
    keywords_negative_sentiment_count = {}

    counter = 1
    for request_item in all_requests:
        try:
            counter = counter + 1
            sentiment_result = request_item['sentiment']
            for keyword in list(request_item['result']):

                keyword_string = keyword

                try:
                    if keyword_string in keywords_count:
                        keywords_count[keyword_string] = keywords_count[keyword_string] + 1 
                    else:
                        keywords_count[keyword_string] = 1
                except Exception as e:
                    print(e)
                    
                if (sentiment_result == 'Positive'):
                    try:
                        if keyword_string in keywords_positive_sentiment_count:
                            keywords_positive_sentiment_count[keyword_string] = keywords_positive_sentiment_count[keyword_string] + 1 
                        else:
                            keywords_positive_sentiment_count[keyword_string] = 1
                    except Exception as e:
                        print(e)
                
                elif (sentiment_result == 'Neutral'):
                    try:
                        if keyword_string in keywords_neutral_sentiment_count:
                            keywords_neutral_sentiment_count[keyword_string] = keywords_neutral_sentiment_count[keyword_string] + 1 
                        else:
                            keywords_neutral_sentiment_count[keyword_string] = 1
                    except Exception as e:
                        print(e)

                elif (sentiment_result == 'Negative'):
                    try:
                        if keyword_string in keywords_negative_sentiment_count:
                            keywords_negative_sentiment_count[keyword_string] = keywords_negative_sentiment_count[keyword_string] + 1 
                        else:
                            keywords_negative_sentiment_count[keyword_string] = 1
                    except Exception as e:
                        print(e)
                    
        except Exception as e:

            print(e)
            continue

    # print(colored(str(entities_count) , 'green'))
    # print(colored(str(entities_positive_sentiment_count) , 'green'))
    # print(colored(str(entities_neutral_sentiment_count) , 'green'))
    # print(colored(str(entities_negative_sentiment_count) , 'green'))

    sortedDict_count = sorted(keywords_count.items(), key=lambda x:x[1], reverse=True)

    # if len(sortedDict_count) > 10:
    #     sortedDict_count = sortedDict_count[0:10]

    data_out_count = []
    data_out_label = []
    data_out_sentiment_positive = []
    data_out_sentiment_neutral = []
    data_out_sentiment_negative = []

    
    for sorted_item in sortedDict_count:
        data_out_count.append(keywords_count[sorted_item[0]])
        data_out_label.append(sorted_item[0])
        
        if (len(data_out_label) == 10):
            break
        
        # Positive
        if (sorted_item[0] in keywords_positive_sentiment_count):
            data_out_sentiment_positive.append(keywords_positive_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_positive.append(0)
        
        # Neutral
        if (sorted_item[0] in keywords_neutral_sentiment_count):
            data_out_sentiment_neutral.append(keywords_neutral_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_neutral.append(0)
        
        # Negative
        if (sorted_item[0] in keywords_negative_sentiment_count):
            data_out_sentiment_negative.append(keywords_negative_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_negative.append(0)
        

    keyword_bulk_list_chart_top_count = {
        "data_positive" :data_out_sentiment_positive,
        "data_neutral" :data_out_sentiment_neutral,
        "data_negative" :data_out_sentiment_negative, 
        "data_count" : data_out_count,
        "labels": data_out_label
    }

    # wordcloud data

    data_out_wordcloud = []
 
    for sorted_item in sortedDict_count:
        if (keywords_count[sorted_item[0]] > 1):
            data_out_wordcloud.append({'text': sorted_item[0] , 'count': keywords_count[sorted_item[0]]})
            # data_out_wordcloud.append(keywords_count[sorted_item[0]])
            # data_out_label_wordcloud.append(sorted_item[0])


    return render(request, 'home/requests-keyword_analytics_report.html', {"msg": 'SUCCESS',
                                                            "segment": 'keyword-analytics',
                                                            "keyword_bulk_list_chart_top_count" : keyword_bulk_list_chart_top_count,
                                                            "keyword_bulk_list_wordcloud" : data_out_wordcloud,
                                                            "bulk_id" : bulk_id,
                                                            "request_list": top_ten_requests})

@login_required(login_url="/login/")
def update_bulk_ner_analytics_ajax_client(request):
    current_pagination = 1   
    try:
        current_pagination = int (request.path.split('/P')[-1])
    except Exception as e:
        pass
    start = (current_pagination-1) * ROW_LIST_SHOW_COUNT
    row_list_bulk, total_count_bulk = get_bulk_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, None, Request_Type.NAMED_ENTITY_RECOGNITION)

    if request.is_ajax and request.method == "GET":
        return render(request, 'includes/bulk_ner_analytics_table_client.html', {'bulk_list':row_list_bulk})

@login_required(login_url="/login/")
def update_bulk_category_youtube_analytics_ajax_client(request):
    current_pagination = 1   
    try:
        current_pagination = int (request.path.split('/P')[-1])
    except Exception as e:
        pass
    start = (current_pagination-1) * ROW_LIST_SHOW_COUNT
    row_list_bulk, total_count_bulk = get_bulk_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, None, Request_Type.YOUTUBE_CATEGORY_EXTRACTION)

    if request.is_ajax and request.method == "GET":
        return render(request, 'includes/bulk_category_youtube_analytics_table_client.html', {'bulk_list':row_list_bulk})

@login_required(login_url="/login/")
def update_bulk_keyword_analytics_ajax_client(request):
    current_pagination = 1   
    try:
        current_pagination = int (request.path.split('/P')[-1])
    except Exception as e:
        pass
    start = (current_pagination-1) * ROW_LIST_SHOW_COUNT
    row_list_bulk, total_count_bulk = get_bulk_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, None, Request_Type.KEYWORD_EXTRACTION)

    if request.is_ajax and request.method == "GET":
        return render(request, 'includes/bulk_keyword_analytics_table_client.html', {'bulk_list':row_list_bulk})

@login_required(login_url="/login/")
def update_bulk_sentiment_analytics_ajax_chart_client(request):
    current_pagination = 1   
    try:
        current_pagination = int (request.path.split('/P')[-1])
    except Exception as e:
        pass
    start = (current_pagination-1) * ROW_LIST_SHOW_COUNT
    row_list_bulk, total_count_bulk = get_bulk_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, request.user, None, Request_Type.SENTIMENT_ANALYSIS)
        
    sentiment_bulk_list_chart = { 
        "data_positive" :[],
        "data_neutral" :[],
        "data_negative" :[], 
        "labels": []
    }

    # reversed to have recent values in right side of chart
    for bulk in reversed(row_list_bulk) :
        sentiment_bulk_list_chart["data_positive"].append(bulk["positive_percent"])
        sentiment_bulk_list_chart["data_neutral"].append(bulk["neutral_percent"])
        sentiment_bulk_list_chart["data_negative"].append(bulk["negative_percent"])
        sentiment_bulk_list_chart["labels"].append(bulk["title"])

    # print(colored(sentiment_bulk_list_chart, 'yellow'))
    if request.is_ajax and request.method == "GET":
        return HttpResponse(json.dumps(sentiment_bulk_list_chart),
        content_type="application/json"
    )
        # return render(request, 'includes/bulk_sentiment_analytics_chart_client.html', { "sentiment_bulk_list_chart" : sentiment_bulk_list_chart })

@login_required(login_url="/login/")
def plans_preview_client(request):

    insert_user_log_db (request.user, 'Plans_Preview')

    plan_list_active, total_enable = get_plan_list_page(request, 0, 100, Plan_Status.ENABLE)

    return render(request, "home/purchase-plans_preview_client.html", {"msg": 'SUCCESS',
                                                                "profile_type" : "admin",
                                                                "segment": 'purchase-plans_preview',
                                                                "plan_list": plan_list_active})

@login_required(login_url="/login/")
def transaction_submit(request):

    insert_user_log_db (request.user, 'Transaction_Submit')

    if request.method == "POST": 

        plan_id = 0
        plan_id_string = '0'
        try:
            plan_id_string = request.path.split('/plan')[-1]
        except Exception as e:
            pass

        plan_id = uuid.UUID(plan_id_string)
        plan_selected = Plan.objects.get(id= plan_id)


        plan_title_selected = ''
        plan_price_selected = ''
        plan_discount_selected = ''
        return_url = ''
        cancel_url= ''
        discount_code= ''

        if (plan_selected != None):

            plan_title_selected = plan_selected.name
            plan_discount_selected = plan_selected.discount
            
            plan_price_selected = str(float (plan_selected.price) * float(100.0 - plan_discount_selected)/100.0) # price * discount
            # return_url = 'http://127.0.0.1:8000/transaction_success/?plan=' + plan_id_string
            return_url = 'https://commentsanalytics.com/transaction_success/?plan=' + plan_id_string
            cancel_return = 'https://commentsanalytics.com/transaction_fail/'
            # plan_price_selected = str((plan_selected.price))

            discount_code = str(request.POST.get('discount_code_input_'+ plan_selected.name).strip())
            if (discount_code != ''):
                discount_value = check_discount_code(discount_code)
                print(colored('input discount value : ' + str(discount_value), 'magenta'))

                if (discount_value != 0):

                    # apply discount_code FEEDBACK100 to update remained_count to Starter Plan
                    if (plan_title_selected == 'Starter' and discount_value == 100):

                        try:
                            user_other_fields_obj = User_Other_Fields.objects.filter(user_id=request.user.id)[0]
                            
                            # check use promotion code just once
                            if user_other_fields_obj.promotion_code_used == False :
                                # update remain in User_Other_Fields in db
                                update_remain_count_user(request, int(plan_selected.count))
                                # update plan in User_Other_Fields in db
                                update_plan_user(request, plan_selected, True)
                                return redirect('/dashboard_client/')
                            else:
                                discount_value = 0
                        except Exception as e:
                            print(colored(str(e), 'red'))
                            discount_value = 0
                        
                    elif (plan_title_selected != 'Starter' and discount_value == 100):
                        # Do not apply discount code 100 Percent for other plans
                        discount_value = 0


                    # apply discount code value input 
                    plan_price_selected = str(float (plan_price_selected) * float(100.0 - discount_value)/100.0) # price * discount_value

        return render(request, "home/purchase-payment.html", {'amount' : plan_price_selected, 'return_url': return_url, 'cencel_url' : cancel_return, 'plan': plan_selected })
    else:
        return redirect('/dashboard_client/')

    #return redirect('https://www.paypal.com/cgi-bin/webscr?cmd=_xclick&business=mahamedani.freelancer@gmail.com&currency=USD&item_name=' + plan_title_selected + '&quantity=1&amount=' + plan_price_selected + '&return=https://commentsanalytics.com/transaction_success/?plan=' + plan_id_string +'&cancel_return=https://commentsanalytics.com/transaction_fail/')
    # return redirect('https://www.sandbox.paypal.com/cgi-bin/webscr?cmd=_xclick&business=mahamedani.freelancer@gmail.com&currency=USD&item_name=' + plan_title_selected + '&quantity=1&amount=' + plan_price_selected + '&return=https://commentsanalytics.com/transaction_success/?plan=' + plan_id_string +'&cancel_return=https://commentsanalytics.com/transaction_fail/')

@login_required(login_url="/login/")
def transaction_success(request):

    insert_user_log_db (request.user, 'Transaction_Success')

    if request.method == "GET": 

        full_url = request.build_absolute_uri()

        paying_id = 0
        try:
            paying_id = full_url.split('&PayerID=')[-1]
        except Exception as e:
            pass
        
        plan_id = 0
        plan_id_string = '0'
        try:
            plan_id_string = full_url.split('/?plan=')[-1].split('&PayerID=')[0]

        except Exception as e:
            pass

        
        plan_id = uuid.UUID(plan_id_string)
        plan_selected = Plan.objects.get(id= plan_id)
        plan_price_selected = 0
        plan_discount_selected = 0.0

        if (plan_selected != None):
            # plan_price_selected = str(plan_selected.price)
            plan_discount_selected = float(plan_selected.discount)

            plan_price_selected = str(float (plan_selected.price) * float(100.0 - plan_discount_selected)/100.0) # price * discount

            plan_selected = Plan.objects.get(id= plan_id)
            try:
                # save Payment record in db
                insert_payment_db (request, Payment_Status.SUCCESS.value, Payment_Type.PAYPALL.value, plan_price_selected, plan_selected, paying_id)
                # update remain in User_Other_Fields in db
                update_remain_count_user(request, int(plan_selected.count))
                # update plan in User_Other_Fields in db
                update_plan_user(request, plan_selected, False)

            except Exception as e:
                print(e)
    return redirect('/transactions_client/')

@login_required(login_url="/login/")
def transaction_fail(request):

    insert_user_log_db (request.user, 'Transaction_Failed')

    if request.method == "POST": 

        paying_id = 0
        try:
            paying_id = request.path.split('?PayerID=')[-1]
        except Exception as e:
            pass

        # save Transaction record in db

    return redirect('/transactions_client/')

@login_required(login_url="/login/")
def transactions_client(request):

    try:
        insert_user_log_db (request.user, 'Transaction_List')

        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT


        row_list_transaction, total_count_transactions, total_earning, monthly_avg_earning, montly_earning_list = get_transaction_list_page(request, start, ROW_LIST_SHOW_COUNT, request.user, Payment_Status.SUCCESS)

        last_pagination = 1
        if (total_count_transactions == 0):
            last_pagination = 1
        elif (total_count_transactions % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_transactions / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_transactions / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/purchase-transactions_client.html", {"msg": 'SUCCESS',
                                                                "segment": 'purchase-transactions_client', 
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_transaction),
                                                                "total_items": total_count_transactions,
                                                                "transaction_list": row_list_transaction})

@login_required(login_url="/login/")
def export_bulk_sentiment_xlsx(request):

    bulk_id = 0
    bulk_id_string = '0'
    page_string = '1'

    try:
        bulk_id_string = request.path.split('/bulk')[-1]
        page_string = request.path.split('/P')[-1].split('/')[0]
    except Exception as e:
        pass
    
    # bulk_id = uuid.UUID(bulk_id_string)
    # bulk_selected = Bulk.objects.get(id= bulk_id)

    query_bulk_id = {'_id' : ObjectId(bulk_id_string)}
    records_bulks = collection_Bulks.find(query_bulk_id)
    bulk_id = bulk_id_string

    if (records_bulks.count() == 0):
        return

    bulk_selected = records_bulks[0]

    excel_file_name = bulk_selected['name'] + '_Result_Sentiment-{date:%Y-%m-%d}'.format( date=datetime.datetime.now() )
    all_requests = get_request_list_bulk_export_mongodb(request, bulk_id)

    # create .xlsx file in response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = "attachment; filename=" + excel_file_name + ".xlsx"

    book = Workbook(response, {'in_memory': True})
    sheet = book.add_worksheet('Sheet1')

    sheet.write(0,0, 'text')
    sheet.write(0,1, 'request_type')
    sheet.write(0,2, 'result')
    sheet.write(0,3, 'time')

    export_time = str(datetime.datetime.now()).split('.')[0]
    row =1
    for request in all_requests:
        sheet.write(row, 0, request['business_name'])
        sheet.write(row, 1, request['request_type'])
        sheet.write(row, 2, request['result'])
        sheet.write(row, 3, export_time)
        row += 1

    # add total_count, positive_sentiment, neutral_sentiment, negative_sentiment
    sheet.write(row, 0, '')
    sheet.write(row + 1, 0, '')
    sheet.write(row + 2, 0, '')

    query_positive = {'bulk': bulk_id, 'result': 'Positive'}
    query_neutral = {'bulk': bulk_id, 'result': 'Neutral'}
    query_negative = {'bulk': bulk_id, 'result': 'Negative'}
    
    positive_count = collection_Requests.find(query_positive).count()
    neutral_count = collection_Requests.find(query_neutral).count()
    negative_count = collection_Requests.find(query_negative).count()
    
    positive_percent = 0
    neutral_percent = 0
    negative_percent = 0
    
    try:
        
        positive_percent = int(positive_count / bulk_selected['total_count'] * 100)
        neutral_percent = int(neutral_count / bulk_selected['total_count'] * 100)
        negative_percent = int(negative_count / bulk_selected['total_count'] * 100)
    except Exception as e:
        print(e)


    sheet.write(row + 3, 0,'positive_sentiment : ' + str(positive_count) + ' ( ' + str(positive_percent) + ' % )')
    sheet.write(row + 4, 0,'neutral_sentiment : ' + str(neutral_count) + ' ( ' + str(neutral_percent) + ' % )')
    sheet.write(row + 5, 0,'negative_sentiment : ' + str(negative_count) + ' ( ' + str(negative_percent) + ' % )')
    sheet.write(row + 6, 0,'total_count : ' + str(bulk_selected['total_count']))

    book.close()

    
    return response
    # return redirect('/requests_bulk_status_client/P' + page_string)

@login_required(login_url="/login/")
def export_bulk_category_xlsx(request):

    bulk_id = 0
    bulk_id_string = '0'
    page_string = '1'

    try:
        bulk_id_string = request.path.split('/bulk')[-1]
        page_string = request.path.split('/P')[-1].split('/')[0]
    except Exception as e:
        pass
    
    # bulk_id = uuid.UUID(bulk_id_string)
    # bulk_selected = Bulk.objects.get(id= bulk_id)

    query_bulk_id = {'_id' : ObjectId(bulk_id_string)}
    records_bulks = collection_Bulks.find(query_bulk_id)
    bulk_id = bulk_id_string

    if (records_bulks.count() == 0):
        return

    bulk_selected = records_bulks[0]

    excel_file_name = bulk_selected['name'] + '_Result_Category-{date:%Y-%m-%d}'.format( date=datetime.datetime.now() )
    all_requests = get_request_list_bulk_export_mongodb(request, bulk_id)

    # create .xlsx file in response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = "attachment; filename=" + excel_file_name + ".xlsx"

    book = Workbook(response, {'in_memory': True})
    sheet = book.add_worksheet('Sheet1')

    sheet.write(0,0, 'text')
    sheet.write(0,1, 'request_type')
    sheet.write(0,2, 'result')
    sheet.write(0,3, 'sentiment')
    sheet.write(0,4, 'time')

    export_time = str(datetime.datetime.now()).split('.')[0]
    row =1
    for request in all_requests:
        result_sentiment = ''
        if ('result_sentiment' in request):
            result_sentiment = request['result_sentiment']
        result = str(request['result'])
        if ('[\'' in result):
            result = result.replace('[\'', '')
        if ('\']' in result):
            result = result.replace('\']', '')

        sheet.write(row, 0, request['business_name'])
        sheet.write(row, 1, request['request_type'])
        sheet.write(row, 2, result)
        sheet.write(row, 3, result_sentiment)
        sheet.write(row, 4, export_time)
        row += 1

  
    book.close()

    
    return response
    # return redirect('/requests_bulk_status_client/P' + page_string)

@login_required(login_url="/login/")
def export_bulk_ner_xlsx(request):

    bulk_id = 0
    bulk_id_string = '0'
    page_string = '1'

    try:
        bulk_id_string = request.path.split('/bulk')[-1]
        page_string = request.path.split('/P')[-1].split('/')[0]
    except Exception as e:
        pass
    
    # bulk_id = uuid.UUID(bulk_id_string)
    # bulk_selected = Bulk.objects.get(id= bulk_id)

    query_bulk_id = {'_id' : ObjectId(bulk_id_string)}
    records_bulks = collection_Bulks.find(query_bulk_id)
    bulk_id = bulk_id_string

    if (records_bulks.count() == 0):
        return

    bulk_selected = records_bulks[0]

    excel_file_name = bulk_selected['name'] + '_Result_NER-{date:%Y-%m-%d}'.format( date=datetime.datetime.now() )
    all_requests = get_request_list_bulk_export_mongodb(request, bulk_id)

    # create .xlsx file in response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = "attachment; filename=" + excel_file_name + ".xlsx"

    book = Workbook(response, {'in_memory': True})
    sheet = book.add_worksheet('Sheet1')

    sheet.write(0,0, 'named_entity')
    sheet.write(0,1, 'count')
    sheet.write(0,2, 'positive_sentiment in Entity')
    sheet.write(0,3, 'neutral_sentiment in Entity')
    sheet.write(0,4, 'negative_sentiment in Entity')
    sheet.write(0,5, 'time')

    export_time = str(datetime.datetime.now()).split('.')[0]
    row =1

    all_requests = get_request_list_bulk_mongodb(request, bulk_id)
    
    entities_count = {}
    entities_positive_sentiment_count = {}
    entities_neutral_sentiment_count = {}
    entities_negative_sentiment_count = {}

    counter = 1
    for request_item in all_requests:
        try:
            counter = counter + 1
            sentiment_result = request_item['sentiment']
            for entity in list(request_item['result']):
                if (len(entity[0]) < 2):
                    continue

                entity_string = entity[0] + ' (' + entity[1] + ')'

                try:
                    if entity_string in entities_count:
                        entities_count[entity_string] = entities_count[entity_string] + 1 
                    else:
                        entities_count[entity_string] = 1
                except Exception as e:
                    print(e)
                    
                if (sentiment_result == 'Positive'):
                    try:
                        if entity_string in entities_positive_sentiment_count:
                            entities_positive_sentiment_count[entity_string] = entities_positive_sentiment_count[entity_string] + 1 
                        else:
                            entities_positive_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)
                
                elif (sentiment_result == 'Neutral'):
                    try:
                        if entity_string in entities_neutral_sentiment_count:
                            entities_neutral_sentiment_count[entity_string] = entities_neutral_sentiment_count[entity_string] + 1 
                        else:
                            entities_neutral_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)

                elif (sentiment_result == 'Negative'):
                    try:
                        if entity_string in entities_negative_sentiment_count:
                            entities_negative_sentiment_count[entity_string] = entities_negative_sentiment_count[entity_string] + 1 
                        else:
                            entities_negative_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)
                    
        except Exception as e:

            print(e)
            continue
    sortedDict_count = sorted(entities_count.items(), key=lambda x:x[1], reverse=True)

    data_out_count = []
    data_out_label = []
    data_out_sentiment_positive = []
    data_out_sentiment_neutral = []
    data_out_sentiment_negative = []

    
    for sorted_item in sortedDict_count:
        data_out_count.append(entities_count[sorted_item[0]])
        data_out_label.append(sorted_item[0])

        # Positive
        if (sorted_item[0] in entities_positive_sentiment_count):
            data_out_sentiment_positive.append(entities_positive_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_positive.append(0)
        
        # Neutral
        if (sorted_item[0] in entities_neutral_sentiment_count):
            data_out_sentiment_neutral.append(entities_neutral_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_neutral.append(0)
        
        # Negative
        if (sorted_item[0] in entities_negative_sentiment_count):
            data_out_sentiment_negative.append(entities_negative_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_negative.append(0)
        

    export_time = str(datetime.datetime.now()).split('.')[0]


    for index_entity, entity in enumerate(data_out_label):
        sheet.write(row, 0, data_out_label[index_entity])
        sheet.write(row, 1, data_out_count[index_entity])
        sheet.write(row, 2, str(data_out_sentiment_positive[index_entity]) + ' (' + str(int(data_out_sentiment_positive[index_entity] / data_out_count[index_entity]* 100)) + ' % )')
        sheet.write(row, 3, str(data_out_sentiment_neutral[index_entity]) + ' (' + str(int(data_out_sentiment_neutral[index_entity] / data_out_count[index_entity]* 100)) + ' % )')
        sheet.write(row, 4, str(data_out_sentiment_negative[index_entity]) + ' (' + str(int(data_out_sentiment_negative[index_entity] / data_out_count[index_entity] * 100)) + ' % )')
        sheet.write(row, 5, export_time)
        row += 1

    book.close()

    
    return response
    # return redirect('/requests_bulk_status_client/P' + page_string)

@login_required(login_url="/login/")
def export_bulk_keyword_xlsx(request):

    bulk_id = 0
    bulk_id_string = '0'
    page_string = '1'

    try:
        bulk_id_string = request.path.split('/bulk')[-1]
        page_string = request.path.split('/P')[-1].split('/')[0]
    except Exception as e:
        pass
    
    # bulk_id = uuid.UUID(bulk_id_string)
    # bulk_selected = Bulk.objects.get(id= bulk_id)

    query_bulk_id = {'_id' : ObjectId(bulk_id_string)}
    records_bulks = collection_Bulks.find(query_bulk_id)
    bulk_id = bulk_id_string

    if (records_bulks.count() == 0):
        return

    bulk_selected = records_bulks[0]

    excel_file_name = bulk_selected['name'] + '_Result_Keyword-{date:%Y-%m-%d}'.format( date=datetime.datetime.now() )
    all_requests = get_request_list_bulk_export_mongodb(request, bulk_id)

    # create .xlsx file in response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = "attachment; filename=" + excel_file_name + ".xlsx"

    book = Workbook(response, {'in_memory': True})
    sheet = book.add_worksheet('Sheet1')

    sheet.write(0,0, 'Keyword')
    sheet.write(0,1, 'count')
    sheet.write(0,2, 'positive_sentiment in Keyword')
    sheet.write(0,3, 'neutral_sentiment in Keyword')
    sheet.write(0,4, 'negative_sentiment in Keyword')
    sheet.write(0,5, 'time')

    export_time = str(datetime.datetime.now()).split('.')[0]
    row =1

    all_requests = get_request_list_bulk_mongodb(request, bulk_id)
    
    entities_count = {}
    entities_positive_sentiment_count = {}
    entities_neutral_sentiment_count = {}
    entities_negative_sentiment_count = {}

    counter = 1
    for request_item in all_requests:
        try:
            counter = counter + 1
            sentiment_result = request_item['sentiment']
            for entity in list(request_item['result']):
                if (len(entity) < 2):
                    continue

                entity_string = entity

                try:
                    if entity_string in entities_count:
                        entities_count[entity_string] = entities_count[entity_string] + 1 
                    else:
                        entities_count[entity_string] = 1
                except Exception as e:
                    print(e)
                    
                if (sentiment_result == 'Positive'):
                    try:
                        if entity_string in entities_positive_sentiment_count:
                            entities_positive_sentiment_count[entity_string] = entities_positive_sentiment_count[entity_string] + 1 
                        else:
                            entities_positive_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)
                
                elif (sentiment_result == 'Neutral'):
                    try:
                        if entity_string in entities_neutral_sentiment_count:
                            entities_neutral_sentiment_count[entity_string] = entities_neutral_sentiment_count[entity_string] + 1 
                        else:
                            entities_neutral_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)

                elif (sentiment_result == 'Negative'):
                    try:
                        if entity_string in entities_negative_sentiment_count:
                            entities_negative_sentiment_count[entity_string] = entities_negative_sentiment_count[entity_string] + 1 
                        else:
                            entities_negative_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)
                    
        except Exception as e:

            print(e)
            continue
    sortedDict_count = sorted(entities_count.items(), key=lambda x:x[1], reverse=True)

    data_out_count = []
    data_out_label = []
    data_out_sentiment_positive = []
    data_out_sentiment_neutral = []
    data_out_sentiment_negative = []

    
    for sorted_item in sortedDict_count:
        data_out_count.append(entities_count[sorted_item[0]])
        data_out_label.append(sorted_item[0])

        # Positive
        if (sorted_item[0] in entities_positive_sentiment_count):
            data_out_sentiment_positive.append(entities_positive_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_positive.append(0)
        
        # Neutral
        if (sorted_item[0] in entities_neutral_sentiment_count):
            data_out_sentiment_neutral.append(entities_neutral_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_neutral.append(0)
        
        # Negative
        if (sorted_item[0] in entities_negative_sentiment_count):
            data_out_sentiment_negative.append(entities_negative_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_negative.append(0)
        

    export_time = str(datetime.datetime.now()).split('.')[0]


    for index_entity, entity in enumerate(data_out_label):
        sheet.write(row, 0, data_out_label[index_entity])
        sheet.write(row, 1, data_out_count[index_entity])
        sheet.write(row, 2, str(data_out_sentiment_positive[index_entity]) + ' (' + str(int(data_out_sentiment_positive[index_entity] / data_out_count[index_entity]* 100)) + ' % )')
        sheet.write(row, 3, str(data_out_sentiment_neutral[index_entity]) + ' (' + str(int(data_out_sentiment_neutral[index_entity] / data_out_count[index_entity]* 100)) + ' % )')
        sheet.write(row, 4, str(data_out_sentiment_negative[index_entity]) + ' (' + str(int(data_out_sentiment_negative[index_entity] / data_out_count[index_entity] * 100)) + ' % )')
        sheet.write(row, 5, export_time)
        row += 1

    book.close()

    
    return response
    # return redirect('/requests_bulk_status_client/P' + page_string)
        
@login_required(login_url="/login/")
def export_bulk_sentiment_csv(request):

    bulk_id = 0
    bulk_id_string = '0'
    page_string = '1'

    try:
        bulk_id_string = request.path.split('/bulk')[-1]
        page_string = request.path.split('/P')[-1].split('/')[0]
    except Exception as e:
        pass
    
    # bulk_id = uuid.UUID(bulk_id_string)
    # bulk_selected = Bulk.objects.get(id= bulk_id)

    query_bulk_id = {'_id' : ObjectId(bulk_id_string)}
    records_bulks = collection_Bulks.find(query_bulk_id)
    bulk_id = bulk_id_string
    bulk_count = records_bulks.count()

    if (bulk_count == 0):
        return

    bulk_selected = records_bulks[0]

    excel_file_name = bulk_selected['name'] + '_Result_Sentiment-{date:%Y-%m-%d}'.format( date=datetime.datetime.now() )
    all_requests = get_request_list_bulk_export_mongodb(request, bulk_id)

    # create .csv file in response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="' + excel_file_name + '.csv"' 
    writer = csv.writer(response)
    writer.writerow(['text','request_type', 'result', 'time'])

    export_time = str(datetime.datetime.now()).split('.')[0]
    for request in all_requests:
        writer.writerow([request['business_name'], request['request_type'], request['result'], export_time])
    

    # add total_count, positive_sentiment, neutral_sentiment, negative_sentiment
    writer.writerow([''])
    writer.writerow([''])
    writer.writerow([''])

    query_positive = {'bulk': bulk_id, 'result': 'Positive'}
    query_neutral = {'bulk': bulk_id, 'result': 'Neutral'}
    query_negative = {'bulk': bulk_id, 'result': 'Negative'}
    
    positive_count = collection_Requests.find(query_positive).count()
    neutral_count = collection_Requests.find(query_neutral).count()
    negative_count = collection_Requests.find(query_negative).count()
    
    positive_percent = 0
    neutral_percent = 0
    negative_percent = 0
    
    try:
        
        positive_percent = int(positive_count / bulk_selected['total_count'] * 100)
        neutral_percent = int(neutral_count / bulk_selected['total_count'] * 100)
        negative_percent = int(negative_count / bulk_selected['total_count'] * 100)
    except Exception as e:
        print(e)


    writer.writerow(['positive_sentiment : ' + str(positive_count) + ' ( ' + str(positive_percent) + ' % )'])
    writer.writerow(['neutral_sentiment : ' + str(neutral_count) + ' ( ' + str(neutral_percent) + ' % )'])
    writer.writerow(['negative_sentiment : ' + str(negative_count) + ' ( ' + str(negative_percent) + ' % )'])
    writer.writerow(['total_count : ' + str(bulk_selected['total_count'])])

    return response
    # return redirect('/requests_bulk_status_client/P' + page_string)

@login_required(login_url="/login/")
def export_bulk_category_csv(request):

    bulk_id = 0
    bulk_id_string = '0'
    page_string = '1'

    try:
        bulk_id_string = request.path.split('/bulk')[-1]
        page_string = request.path.split('/P')[-1].split('/')[0]
    except Exception as e:
        pass
    
    # bulk_id = uuid.UUID(bulk_id_string)
    # bulk_selected = Bulk.objects.get(id= bulk_id)

    query_bulk_id = {'_id' : ObjectId(bulk_id_string)}
    records_bulks = collection_Bulks.find(query_bulk_id)
    bulk_id = bulk_id_string
    bulk_count = records_bulks.count()

    if (bulk_count == 0):
        return

    bulk_selected = records_bulks[0]

    excel_file_name = bulk_selected['name'] + '_Result_Category-{date:%Y-%m-%d}'.format( date=datetime.datetime.now() )
    all_requests = get_request_list_bulk_export_mongodb(request, bulk_id)

    # create .csv file in response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="' + excel_file_name + '.csv"' 
    writer = csv.writer(response)
    writer.writerow(['text','request_type', 'category', 'sentiment', 'time'])

    export_time = str(datetime.datetime.now()).split('.')[0]
    for request in all_requests:
        result_sentiment = ''
        if ('result_sentiment' in request):
            result_sentiment = request['result_sentiment']
        result = str(request['result'])
        if ('[\'' in result):
            result = result.replace('[\'', '')
        if ('\']' in result):
            result = result.replace('\']', '')
        writer.writerow([request['business_name'], request['request_type'], result, result_sentiment, export_time])
    

    
    return response
    # return redirect('/requests_bulk_status_client/P' + page_string)

@login_required(login_url="/login/")
def export_bulk_ner_csv(request):

    bulk_id = 0
    bulk_id_string = '0'
    page_string = '1'

    try:
        bulk_id_string = request.path.split('/bulk')[-1]
        page_string = request.path.split('/P')[-1].split('/')[0]
    except Exception as e:
        pass
    
    # bulk_id = uuid.UUID(bulk_id_string)
    # bulk_selected = Bulk.objects.get(id= bulk_id)

    query_bulk_id = {'_id' : ObjectId(bulk_id_string)}
    records_bulks = collection_Bulks.find(query_bulk_id)
    bulk_id = bulk_id_string
    bulk_count = records_bulks.count()

    if (bulk_count == 0):
        return

    bulk_selected = records_bulks[0]

    excel_file_name = bulk_selected['name'] + '_Result_NER-{date:%Y-%m-%d}'.format( date=datetime.datetime.now() )
    all_requests = get_request_list_bulk_mongodb(request, bulk_id)
    
    entities_count = {}
    entities_positive_sentiment_count = {}
    entities_neutral_sentiment_count = {}
    entities_negative_sentiment_count = {}

    counter = 1
    for request_item in all_requests:
        try:
            counter = counter + 1
            sentiment_result = request_item['sentiment']
            for entity in list(request_item['result']):
                if (len(entity[0]) < 2):
                    continue

                entity_string = entity[0] + ' (' + entity[1] + ')'

                try:
                    if entity_string in entities_count:
                        entities_count[entity_string] = entities_count[entity_string] + 1 
                    else:
                        entities_count[entity_string] = 1
                except Exception as e:
                    print(e)
                    
                if (sentiment_result == 'Positive'):
                    try:
                        if entity_string in entities_positive_sentiment_count:
                            entities_positive_sentiment_count[entity_string] = entities_positive_sentiment_count[entity_string] + 1 
                        else:
                            entities_positive_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)
                
                elif (sentiment_result == 'Neutral'):
                    try:
                        if entity_string in entities_neutral_sentiment_count:
                            entities_neutral_sentiment_count[entity_string] = entities_neutral_sentiment_count[entity_string] + 1 
                        else:
                            entities_neutral_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)

                elif (sentiment_result == 'Negative'):
                    try:
                        if entity_string in entities_negative_sentiment_count:
                            entities_negative_sentiment_count[entity_string] = entities_negative_sentiment_count[entity_string] + 1 
                        else:
                            entities_negative_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)
                    
        except Exception as e:

            print(e)
            continue
    sortedDict_count = sorted(entities_count.items(), key=lambda x:x[1], reverse=True)

    data_out_count = []
    data_out_label = []
    data_out_sentiment_positive = []
    data_out_sentiment_neutral = []
    data_out_sentiment_negative = []

    
    for sorted_item in sortedDict_count:
        data_out_count.append(entities_count[sorted_item[0]])
        data_out_label.append(sorted_item[0])

        # Positive
        if (sorted_item[0] in entities_positive_sentiment_count):
            data_out_sentiment_positive.append(entities_positive_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_positive.append(0)
        
        # Neutral
        if (sorted_item[0] in entities_neutral_sentiment_count):
            data_out_sentiment_neutral.append(entities_neutral_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_neutral.append(0)
        
        # Negative
        if (sorted_item[0] in entities_negative_sentiment_count):
            data_out_sentiment_negative.append(entities_negative_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_negative.append(0)
        

    # create .csv file in response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="' + excel_file_name + '.csv"' 
    writer = csv.writer(response)
    writer.writerow(['named_entity','count', 'positive_sentiment in Entity', 'neutral_sentiment in Entity', 'negative_sentiment in Entity', 'time'])

    export_time = str(datetime.datetime.now()).split('.')[0]
    for index_entity, entity in enumerate(data_out_label):
        writer.writerow([data_out_label[index_entity], data_out_count[index_entity], 
        str(data_out_sentiment_positive[index_entity]) + ' (' + str(int(data_out_sentiment_positive[index_entity] / data_out_count[index_entity]* 100)) + ' % )',
        str(data_out_sentiment_neutral[index_entity]) + ' (' + str(int(data_out_sentiment_neutral[index_entity] / data_out_count[index_entity]* 100)) + ' % )',
        str(data_out_sentiment_negative[index_entity]) + ' (' + str(int(data_out_sentiment_negative[index_entity] / data_out_count[index_entity] * 100)) + ' % )'
        , export_time])
    

    
    return response
    # return redirect('/requests_bulk_status_client/P' + page_string)

@login_required(login_url="/login/")
def export_bulk_keyword_csv(request):

    bulk_id = 0
    bulk_id_string = '0'
    page_string = '1'

    try:
        bulk_id_string = request.path.split('/bulk')[-1]
        page_string = request.path.split('/P')[-1].split('/')[0]
    except Exception as e:
        pass
    
    # bulk_id = uuid.UUID(bulk_id_string)
    # bulk_selected = Bulk.objects.get(id= bulk_id)

    query_bulk_id = {'_id' : ObjectId(bulk_id_string)}
    records_bulks = collection_Bulks.find(query_bulk_id)
    bulk_id = bulk_id_string
    bulk_count = records_bulks.count()

    if (bulk_count == 0):
        return

    bulk_selected = records_bulks[0]

    excel_file_name = bulk_selected['name'] + '_Result_Keyword-{date:%Y-%m-%d}'.format( date=datetime.datetime.now() )
    all_requests = get_request_list_bulk_mongodb(request, bulk_id)

    entities_count = {}
    entities_positive_sentiment_count = {}
    entities_neutral_sentiment_count = {}
    entities_negative_sentiment_count = {}

    counter = 1
    for request_item in all_requests:
        try:
            counter = counter + 1
            sentiment_result = request_item['sentiment']
            for entity in list(request_item['result']):
                if (len(entity) < 2):
                    continue

                entity_string = entity

                try:
                    if entity_string in entities_count:
                        entities_count[entity_string] = entities_count[entity_string] + 1 
                    else:
                        entities_count[entity_string] = 1
                except Exception as e:
                    print(e)
                    
                if (sentiment_result == 'Positive'):
                    try:
                        if entity_string in entities_positive_sentiment_count:
                            entities_positive_sentiment_count[entity_string] = entities_positive_sentiment_count[entity_string] + 1 
                        else:
                            entities_positive_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)
                
                elif (sentiment_result == 'Neutral'):
                    try:
                        if entity_string in entities_neutral_sentiment_count:
                            entities_neutral_sentiment_count[entity_string] = entities_neutral_sentiment_count[entity_string] + 1 
                        else:
                            entities_neutral_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)

                elif (sentiment_result == 'Negative'):
                    try:
                        if entity_string in entities_negative_sentiment_count:
                            entities_negative_sentiment_count[entity_string] = entities_negative_sentiment_count[entity_string] + 1 
                        else:
                            entities_negative_sentiment_count[entity_string] = 1
                    except Exception as e:
                        print(e)
                    
        except Exception as e:

            print(e)
            continue
    sortedDict_count = sorted(entities_count.items(), key=lambda x:x[1], reverse=True)

    data_out_count = []
    data_out_label = []
    data_out_sentiment_positive = []
    data_out_sentiment_neutral = []
    data_out_sentiment_negative = []

    
    for sorted_item in sortedDict_count:
        data_out_count.append(entities_count[sorted_item[0]])
        data_out_label.append(sorted_item[0])

        # Positive
        if (sorted_item[0] in entities_positive_sentiment_count):
            data_out_sentiment_positive.append(entities_positive_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_positive.append(0)
        
        # Neutral
        if (sorted_item[0] in entities_neutral_sentiment_count):
            data_out_sentiment_neutral.append(entities_neutral_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_neutral.append(0)
        
        # Negative
        if (sorted_item[0] in entities_negative_sentiment_count):
            data_out_sentiment_negative.append(entities_negative_sentiment_count[sorted_item[0]])
        else:
            data_out_sentiment_negative.append(0)
        

    # create .csv file in response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="' + excel_file_name + '.csv"' 
    writer = csv.writer(response)
    writer.writerow(['Keyword','count', 'positive_sentiment in Keyword', 'neutral_sentiment in Keyword', 'negative_sentiment in Keyword', 'time'])

    export_time = str(datetime.datetime.now()).split('.')[0]
    for index_entity, entity in enumerate(data_out_label):
        writer.writerow([data_out_label[index_entity], data_out_count[index_entity], 
        str(data_out_sentiment_positive[index_entity]) + ' (' + str(int(data_out_sentiment_positive[index_entity] / data_out_count[index_entity]* 100)) + ' % )',
        str(data_out_sentiment_neutral[index_entity]) + ' (' + str(int(data_out_sentiment_neutral[index_entity] / data_out_count[index_entity]* 100)) + ' % )',
        str(data_out_sentiment_negative[index_entity]) + ' (' + str(int(data_out_sentiment_negative[index_entity] / data_out_count[index_entity] * 100)) + ' % )'
        , export_time])
    

    
    return response
    # return redirect('/requests_bulk_status_client/P' + page_string)
                    
@login_required(login_url="/login/")
def invoice(request):

    name = ''
    user_name = ''
    address = ''
    email = ''
    invoice_number = '00000'
    plan_title = 'Free'
    count = 0
    price = 0
    discount = 0
    price_discounted = 0
    payment_date = datetime.datetime.now
    payment_type = 'Paypal'


    try:
        payment_id_string = '0'

        try:
            payment_id_string = request.path.split('invoice/')[-1]
        except Exception as e:
            pass
        
        payment_id = uuid.UUID(payment_id_string)
        payment_selected = Payment.objects.get(id= payment_id)
        user_name = payment_selected.user.username
        name += payment_selected.user.first_name
        name += payment_selected.user.last_name
        email = payment_selected.user.email
        invoice_number = payment_selected.paying_id
        plan_title = payment_selected.plan.name
        count = payment_selected.plan.count
        price = payment_selected.plan.price
        discount = payment_selected.plan.discount
        price_discounted = float(price * float(( 100 - discount )/100))
        discount_value = price - price_discounted
        payment_date = payment_selected.payment_time_slot.date()
        payment_type = payment_selected.payment_type

    except Exception as e:
        print(e)

    return render(request, "home/purchase-invoice.html", {"msg": 'SUCCESS',
                                                          "payment_id" : payment_id_string,
                                                          "segment": 'purchase-transactions_client', 
                                                          "name" : name,
                                                          "user_name" : user_name,
                                                          "email" : email,
                                                          "invoice_number" : invoice_number,
                                                          "plan_title" : plan_title,
                                                          "count" : count,
                                                          "price" : price,
                                                          "price_discounted" : price_discounted,
                                                          "discount_value" : discount_value,
                                                          "discount" : discount,
                                                          "payment_date": payment_date,
                                                          "payment_type": payment_type
                                                             })    
    
@login_required(login_url="/login/")
def download_pdf_invoice(request):

    name = ''
    user_name = ''
    address = ''
    email = ''
    invoice_number = '00000'
    plan_title = 'Free'
    count = 0
    price = 0
    discount = 0
    price_discounted = 0
    payment_date = datetime.datetime.now
    payment_type = 'Paypal'


    try:
        payment_id_string = '0'

        try:
            payment_id_string = request.path.split('invoice_pdf/')[-1]
        except Exception as e:
            pass
        
        payment_id = uuid.UUID(payment_id_string)
        payment_selected = Payment.objects.get(id= payment_id)
        user_name = payment_selected.user.username
        name += payment_selected.user.first_name
        name += payment_selected.user.last_name
        email = payment_selected.user.email
        invoice_number = payment_selected.paying_id
        plan_title = payment_selected.plan.name
        count = payment_selected.plan.count
        price = payment_selected.plan.price
        discount = payment_selected.plan.discount
        price_discounted = float(price * float(( 100 - discount )/100))
        discount_value = price - price_discounted
        payment_date = payment_selected.payment_time_slot.date()
        payment_type = payment_selected.payment_type

    except Exception as e:
        print(e)


    # Create a file-like buffer to receive PDF data.
    buffer = io.BytesIO()

    # Create the PDF object, using the buffer as its "file."
    p = canvas.Canvas(buffer)

    # Draw things on the PDF. Here's where the PDF generation happens.
    # See the ReportLab documentation for the full list of functionality.
    p.setFont('Helvetica-Bold', 12)
    p.drawString(412, 700, "Comments Analytics.")
    p.setFont('Helvetica', 10)
    p.drawString(412, 670, "Lentersweg 36")
    p.drawString(412, 650, "Hamburg, Germany")
    p.setFont('Helvetica-Bold', 10)
    p.drawString(412, 630, "info@commentsanalytics.com")


    p.line(50, 600, 560, 600)

    # logo_name = '~/Documents/Django/BusinessInfo/apps/static/assets/img/brand/light1.svg'
    # a = Image.open("test.png")  
    # fp = open("~/Documents/Projects/socialnetworkDirectory/BusinessInfo/apps/home/test.png","rb")
    # img = PIL.Image.open(fp)
    # img.show()

    # logo = ImageReader('./test.png')
    # p.drawImage(a, 100, 100, mask='auto')


    p.setFont('Helvetica-Bold', 18)
    p.drawString(220, 560, "Invoice")

    p.setFont('Helvetica-Bold', 15)
    p.drawString(300, 560, "Paid")
    p.roundRect(292, 555, 45, 20, 4, stroke=1, fill=0)


    p.setFont('Helvetica-Bold', 10)
    p.drawString(60, 495, "Client Information:")

    p.setFont('Helvetica', 10)
    p.drawString(60, 465, "User_name : " + user_name)
    p.setFont('Helvetica-Bold', 10)
    p.drawString(60, 445, email)


    p.setFont('Helvetica-Bold', 10)
    p.drawString(320, 500, "Invoice No.")
    p.setFont('Helvetica', 10)
    p.drawString(420, 500, invoice_number)
    p.setFont('Helvetica-Bold', 10)
    p.drawString(320, 480, "Date Issued:")
    p.setFont('Helvetica', 10)
    p.drawString(420, 480, str(payment_date))
    p.setFont('Helvetica-Bold', 10)
    p.drawString(320, 460, "Date Due:")
    p.setFont('Helvetica', 10)
    p.drawString(420, 460, str(payment_date))


    p.line(50, 400, 560, 400)

    p.setFont('Helvetica-Bold', 8)
    p.drawString(50, 390, "ITEM")
    p.drawString(162, 390, "DESCRIPTION")
    p.drawString(380, 390, "PRICE")
    p.drawString(430, 390, "QTY")
    p.drawString(450, 390, "TYPE")
    p.drawString(495, 390, "TOTAL")

    p.setFont('Helvetica-Bold', 9)
    p.drawString(50, 375, "Plan : " + plan_title)
    p.setFont('Helvetica', 8)
    p.drawString(162, 375, "Using Comments Analytics services for " + str(count) + " words.")
    p.setFont('Helvetica', 9)
    p.drawString(380, 375, "$ "+ str(price))
    p.drawString(430, 375, "1")
    p.drawString(450, 375, payment_type)
    p.drawString(495, 375, "$ "+ str(price))

    p.line(50, 362, 560, 362)

    p.setFont('Helvetica-Bold', 9)
    p.drawString(425, 330, "Subtotal")
    p.line(420, 320, 560, 320)
    p.drawString(425, 300, "Discount (" + str(discount) + "%)")
    p.line(420, 290, 560, 290)
    p.drawString(425, 270, "VAT (0%)")
    p.line(420, 260, 560, 260)
    p.drawString(425, 240, "Total")
    p.line(420, 230, 560, 230)


    p.setFont('Helvetica', 9)
    p.drawString(495, 330, "$ "+ str(price))
    p.drawString(495, 300, "$ "+ str(discount_value))
    p.drawString(495, 270, "$ "+ str(0))
    p.drawString(495, 240, "$ "+ str(price_discounted))








    # Close the PDF object cleanly, and we're done.
    p.showPage()
    p.save()

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='invoice_' + invoice_number + '.pdf')
   
@login_required(login_url="/login/")
def tickets_client(request):

    try:
        insert_user_log_db (request.user, 'Tickets_page')

        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT


        row_list_ticket, total_count_ticket, unread_count_ticket = get_ticket_list_page(request, start, ROW_LIST_SHOW_COUNT, request.user)

        last_pagination = 1
        if (total_count_ticket == 0):
            last_pagination = 1
        elif (total_count_ticket % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_ticket / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_ticket / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/tickets_client.html", {"msg": 'SUCCESS',
                                                                "segment": 'tickets_client', 
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_ticket),
                                                                "total_items": total_count_ticket,
                                                                "unread_count_ticket" : unread_count_ticket,
                                                                "ticket_list": row_list_ticket})

@login_required(login_url="/login/")
def new_ticket_client(request):

    ticket_id = 0
    try:
        insert_user_log_db (request.user, 'New_Ticket')
        ticket_title= 'Not Inserted'

        if request.method == "POST": 
            ticket_title = request.POST.get('ticket_title').strip()

        ticket_id = insert_ticket_db_client(request.user, ticket_title)
        

    except Exception as e:
        print(e)
        pass

    return redirect ('/ticket_messages_client/ticket_' + str(ticket_id))

@login_required(login_url="/login/")
def ticket_messages_client(request):

    ticket_title = ''
    ticket_message_list = []
    unread_count_ticket = 0

    try:

        insert_user_log_db (request.user, 'Ticket_Messages_page')

        ticket_id = 0
        ticket_id_string = '0'
        try:
            ticket_id_string = request.path.split('/ticket_')[-1]
        except Exception as e:
            pass

        ticket_id = uuid.UUID(ticket_id_string)
        ticket_selected = Ticket.objects.get(id= ticket_id)
        ticket_title = ticket_selected.title

        ticket_message_list, total_count = get_ticket_messages_list(ticket_id)

        # update seen status admin of ticket
        ticket_selected.client_seen_status = Seen_Status.SEEN.value
        ticket_selected.save()

        # update messsage seen status for admin messages
        ticket_message_objects_list = Ticket_Message.objects.all().filter(ticket_id=ticket_id)
        for message in ticket_message_objects_list:
            if (message.user_type == User_Type.ADMINISTRATOR.value):
                try:
                    message.seen_status = Seen_Status.SEEN.value
                    message.save()
                except Exception as e:
                    print(e)

        row_list_ticket, total_count_ticket, unread_count_ticket = get_ticket_list_page(request, 0, ROW_LIST_SHOW_COUNT, request.user)


    except Exception as e:
        print(e)
        pass

    return render(request, "home/ticket_messages_client.html", {"msg": 'SUCCESS',
                                                                "segment": 'tickets_client', 
                                                                "title": ticket_title,
                                                                "ticket_id" : ticket_id_string,
                                                                "unread_count_ticket" : unread_count_ticket,
                                                                "ticket_message_list": ticket_message_list})

@login_required(login_url="/login/")
def new_message_client(request):

    ticket_id = 0
    try:
        insert_user_log_db (request.user, 'New_Message')

        ticket_id = 0
        ticket_id_string = '0'
        try:
            ticket_id_string = request.path.split('/ticket_')[-1]
        except Exception as e:
            pass

        ticket_id = uuid.UUID(ticket_id_string)
        ticket_selected = Ticket.objects.get(id= ticket_id)

        message_body= ''

        if request.method == "POST": 
            message_body = request.POST.get('message_body').strip()

        # insert new message to ticket selected
        insert_message_db(request.user, message_body, ticket_selected, User_Type.CLIENT.value)
        
        # update ticket message count and empty status
        try:
            ticket_selected.message_count = ticket_selected.message_count + 1
            ticket_selected.empty_status = Empty_Status.NOT_EMPTY.value
            ticket_selected.admin_seen_status = Seen_Status.UNSEEN.value
            ticket_selected.question_time_slot = datetime.datetime.now()

            if (ticket_selected.status == Ticket_Status.ADMIN_ANSWERED.value or ticket_selected.status == Ticket_Status.ADMIN_NOTIFICATION.value):
                ticket_selected.status = Ticket_Status.USER_ANSWERED.value

            ticket_selected.save()

            # send email to admin
            send_email_from_client_to_admin(request.user, message_body)

        except Exception as e:
            print(colored(str(e), 'red'))



    except Exception as e:
        print(e)
        pass

    return redirect ('/ticket_messages_client/ticket_' + str(ticket_id_string))

@login_required(login_url="/login/")
def delete_ticket(request):

    ticket_id = 0
    try:
        insert_user_log_db (request.user, 'Delete_Ticket')

        ticket_id = 0
        ticket_id_string = '0'
        try:
            ticket_id_string = request.path.split('/ticket_')[-1]
        except Exception as e:
            pass

        ticket_id = uuid.UUID(ticket_id_string)
        ticket_selected = Ticket.objects.get(id= ticket_id)

        ticket_message_list = Ticket_Message.objects.all().filter(ticket_id=ticket_id)

        # deleting all messages of ticket first
        for message in ticket_message_list:

            try:
                message.delete()
            except Exception as e:
                print(colored('Exception in deleting message : ' + str(e), 'red'))

        try:
            # then delete the ticket object itself
            ticket_selected.delete()
        except Exception as e:
                print(colored('Exception in deleting ticket : ' + str(e), 'red'))



    except Exception as e:
        print(e)
        pass

    return redirect ('/tickets_client/')

@login_required(login_url="/login/")
def delete_message_client(request):

    ticket_id = 0
    try:
        insert_user_log_db (request.user, 'Delete_Message')

        message_id = 0
        message_id_string = '0'
        ticket_id = ''

        try:
            message_id_string = request.path.split('/message_')[-1]
        except Exception as e:
            pass

        message_id = uuid.UUID(message_id_string)
        message_selected = Ticket_Message.objects.get(id= message_id)
        ticket_selected = message_selected.ticket
        ticket_id = ticket_selected.id

        # deleting messages from Ticket_Message first

        try:
            message_selected.delete()
        except Exception as e:
            print(colored('Exception in deleting message : ' + str(e), 'red'))

        # then update the ticket object Empty_Status and message_count and status
        try:
            if (ticket_selected.message_count == 1):
                ticket_selected.empty_status = Empty_Status.EMPTY.value

            ticket_selected.message_count = ticket_selected.message_count - 1
            ticket_selected.save()

        except Exception as e:
                print(colored('Exception in updating ticket : ' + str(e), 'red'))



    except Exception as e:
        print(e)
        pass

    return redirect ('/ticket_messages_client/ticket_' + str(ticket_id))

############################################# Admin ############################################################

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def dashboard_admin(request):

    row_list_ticket, total_count_ticket, unread_count_ticket = get_ticket_list_page(request, 0, ROW_LIST_SHOW_COUNT, None)


    return render(request, "home/dashboard_admin.html", {"msg": 'SUCCESS',
                                                         "total_count_ticket" : total_count_ticket,
                                                         "unread_count_ticket" : unread_count_ticket
                                                        })

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def userslist_admin(request):
    
    current_pagination = 1
       
    try:
        current_pagination = int (request.path.split('/P')[-1])
    except Exception as e:
        pass
    
    start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

    row_list_users, total_count_users = get_user_list_page(request, start, ROW_LIST_SHOW_COUNT)

    total_users, monthly_avg_users, montly_users_list = get_users_statistics()

    last_pagination = 1
    if (total_count_users == 0):
        last_pagination = 1
    elif (total_count_users % ROW_LIST_SHOW_COUNT == 0):
        last_pagination = int (total_count_users / ROW_LIST_SHOW_COUNT)
    else:
        last_pagination = int (total_count_users / ROW_LIST_SHOW_COUNT) + 1


    return render(request, "home/users-list_admin.html", {"msg": 'SUCCESS',
                                                     "segment": 'users-list',
                                                     "users_list": row_list_users,
                                                     "current_pagination": current_pagination,
                                                     "last_pagination": last_pagination,
                                                     "page_limit": len(row_list_users),
                                                     "monthly_users_list" : montly_users_list,
                                                     "total_items": total_count_users})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def requests_bulk_status_admin(request):

    try:
        if request.method == "POST": # just in confirm new bulk data buttton clicked
            pass

        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_bulk, total_count_bulk = get_bulk_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, None) # pass None for admin all users report


        last_pagination = 1
        if (total_count_bulk == 0):
            last_pagination = 1
        elif (total_count_bulk % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_bulk / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_bulk / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/requests-bulk_status_admin.html", {"msg": 'SUCCESS',
                                                                        "segment": 'requests-bulk_status',
                                                                        "current_pagination": current_pagination,
                                                                        "last_pagination": last_pagination,
                                                                        "page_limit": len(row_list_bulk),
                                                                        "total_items": total_count_bulk,
                                                                        "bulk_list": row_list_bulk})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def requests_single_admin(request):

    requests_types_tuple = list(Request_Type.choices())
    requests_types = list((i[1]) for i in requests_types_tuple)
    request_result = ''
    try:
        if request.method == "POST": 
           pass

        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_request, total_count_request = get_request_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, None, Request_Run_Type.SINGLE)

        last_pagination = 1
        if (total_count_request == 0):
            last_pagination = 1
        elif (total_count_request % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/requests-single_admin.html", {"msg": 'SUCCESS',
                                                                "segment": 'requests-single', 
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_request),
                                                                "total_items": total_count_request,
                                                                "request_list": row_list_request})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def requests_bulk_admin(request):

    try:
 
        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_request, total_count_request = get_request_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, None, Request_Run_Type.BULK)

        last_pagination = 1
        if (total_count_request == 0):
            last_pagination = 1
        elif (total_count_request % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/requests-bulk_admin.html", {"msg": 'SUCCESS',
                                                                "segment": 'requests-bulk_admin', 
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_request),
                                                                "total_items": total_count_request,
                                                                "request_list": row_list_request})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def requests_demo_admin(request):

    try:
 
        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_request, total_count_request = get_request_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, None, Request_Run_Type.DEMO)

        last_pagination = 1
        if (total_count_request == 0):
            last_pagination = 1
        elif (total_count_request % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/requests-demo_admin.html", {"msg": 'SUCCESS',
                                                                "segment": 'requests-demo', 
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_request),
                                                                "total_items": total_count_request,
                                                                "request_list": row_list_request})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def plans_admin(request):

    try:
        
        if request.method == "POST": 

            plan_title = request.POST.get('input_plan_title').strip()
            plan_status = request.POST.get('input_plan_status').strip()
            plan_description = request.POST.get('input_plan_description').strip()
            plan_count = 0
            plan_price = 0
            plan_discount = 0
            plan_order = 0


            if (plan_count != ''):
                try:
                    plan_count = int(request.POST.get('input_plan_count').strip())
                except Exception as e:
                    print(e)
                    pass
            if (request.POST.get('input_plan_price').strip() != ''):
                try:
                    plan_price = float(request.POST.get('input_plan_price').strip())
                except Exception as e:
                    print(e)
                    pass
            if (request.POST.get('input_plan_discount').strip() != ''):
                try:
                    plan_discount = int(request.POST.get('input_plan_discount').strip())
                except Exception as e:
                    print(e)
                    pass
            if (request.POST.get('input_plan_order').strip() != ''):
                try:
                    plan_order = int(request.POST.get('input_plan_order').strip())
                except Exception as e:
                    print(e)
                    pass

            plan_tag = request.POST.get('input_plan_tag').strip()

            
            # use hidden variable to distinguish new or edit 
            form_type = request.POST.get('input_form_type').strip()
            plan_id_string = request.POST.get('input_plan_id_selected').strip()

            if (form_type == 'New'):
                insert_plan_db (plan_title, plan_status, plan_description, plan_count, plan_price, plan_discount, plan_order, plan_tag)
            elif (form_type == "Edit"):
                try:
                    plan_id = uuid.UUID(plan_id_string)
                    plan_obj = Plan.objects.get(id= plan_id)
                    plan_obj.name = plan_title
                    plan_obj.status = plan_status
                    plan_obj.description = plan_description
                    plan_obj.count = plan_count
                    plan_obj.price = plan_price
                    plan_obj.discount = plan_discount
                    plan_obj.order = plan_order
                    plan_obj.tag = plan_tag
                    plan_obj.modify_time_slot = datetime.datetime.now()
                    
                    plan_obj.save()
                except Exception as e:
                    print('Exceptionin Edit Plans')
        else:
            pass

        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_plan, total_count_request = get_plan_list_page(request, start, ROW_LIST_SHOW_COUNT, None)

        last_pagination = 1
        if (total_count_request == 0):
            last_pagination = 1
        elif (total_count_request % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_request / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/purchase-plans_admin.html", {"msg": 'SUCCESS',
                                                                "segment": 'purchase-plans_admin', 
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_plan),
                                                                "total_items": total_count_request,
                                                                "plan_list": row_list_plan})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def plans_admin_new_plan(request):

    plan_status_tuple = list(Plan_Status.choices())
    plan_status = list((i[1]) for i in plan_status_tuple) 
    return render(request, "home/purchase-plans_admin_new_plan.html", {"msg": 'SUCCESS',
                                                                "form_type" : "New",
                                                                "segment": 'purchase-plans_admin_new_plan',
                                                                "plan_status": plan_status})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def plans_admin_edit_plan(request):

    plan_status_tuple = list(Plan_Status.choices())
    plan_status = list((i[1]) for i in plan_status_tuple)

    plan_id = 0
    plan_id_string = '0'
    try:
        plan_id_string = request.path.split('/plan')[-1]
    except Exception as e:
        pass

    plan_id = uuid.UUID(plan_id_string)
    plan_selected = Plan.objects.get(id= plan_id)


    plan_title_selected = ''
    plan_status_selected = ''
    plan_description_selected = ''
    plan_count_selected = ''
    plan_price_selected = ''
    plan_discount_selected = ''
    plan_order_selected = ''
    plan_tag_selected = ''

    if (plan_selected != None):

        plan_title_selected = plan_selected.name
        plan_status_selected = plan_selected.status
        plan_description_selected = plan_selected.description
        plan_count_selected = plan_selected.count
        plan_price_selected = plan_selected.price
        plan_discount_selected = plan_selected.discount
        plan_order_selected = plan_selected.order
        plan_tag_selected = plan_selected.tag

    return render(request, "home/purchase-plans_admin_new_plan.html", {"msg": 'SUCCESS',
                                                                "form_type" : "Edit",
                                                                "segment": 'purchase-plans_admin_new_plan',
                                                                "plan_id_selected": plan_id_string,
                                                                "plan_title_selected": plan_title_selected,
                                                                "plan_status_selected": plan_status_selected,
                                                                "plan_description_selected": plan_description_selected,
                                                                "plan_count_selected": plan_count_selected,
                                                                "plan_price_selected": plan_price_selected,
                                                                "plan_discount_selected": plan_discount_selected,
                                                                "plan_order_selected": plan_order_selected,
                                                                "plan_tag_selected": plan_tag_selected,
                                                                "plan_status": plan_status})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def plans_preview_admin(request):

    plan_list_active, total_enable = get_plan_list_page(request, 0, 100, Plan_Status.ENABLE)

    return render(request, "home/purchase-plans_preview_admin.html", {"msg": 'SUCCESS',
                                                                "profile_type" : "admin",
                                                                "segment": 'purchase-plans_preview',
                                                                "plan_list": plan_list_active})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def transactions_admin(request):

    try:

        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_transaction, total_count_transactions, total_earning, monthly_avg_earning, montly_earning_list = get_transaction_list_page(request, start, ROW_LIST_SHOW_COUNT, None, Payment_Status.SUCCESS)
        
        last_pagination = 1
        if (total_count_transactions == 0):
            last_pagination = 1
        elif (total_count_transactions % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_transactions / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_transactions / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/purchase-transactions_admin.html", {"msg": 'SUCCESS',
                                                                "segment": 'purchase-transactions_admin', 
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_transaction),
                                                                "total_items": total_count_transactions,
                                                                "total_earning" : total_earning,
                                                                "monthly_avg_earning" : monthly_avg_earning,
                                                                "montly_earning_list" : montly_earning_list,
                                                                "transaction_list": row_list_transaction})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def user_log_admin(request):

    try:

        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_user_log, total_count_user_log = get_user_log_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, None)
        
        last_pagination = 1
        if (total_count_user_log == 0):
            last_pagination = 1
        elif (total_count_user_log % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_user_log / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_user_log / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/users-log_admin.html", {"msg": 'SUCCESS',
                                                                "segment": 'users-log', 
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_user_log),
                                                                "total_items": total_count_user_log,
                                                                "user_log_list": row_list_user_log})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def tickets_admin(request):

    try:

        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT


        row_list_ticket, total_count_ticket, unread_count_ticket = get_ticket_list_page(request, start, ROW_LIST_SHOW_COUNT, None)
        all_users, total_count_user = get_user_list_page(request,0 , 100)

        last_pagination = 1
        if (total_count_ticket == 0):
            last_pagination = 1
        elif (total_count_ticket % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_ticket / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_ticket / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/tickets_admin.html", {"msg": 'SUCCESS',
                                                                "segment": 'tickets_client', 
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_ticket),
                                                                "total_items": total_count_ticket,
                                                                "unread_count_ticket" : unread_count_ticket,
                                                                "all_users_list" : all_users,
                                                                "ticket_list": row_list_ticket})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def ticket_messages_admin(request):

    ticket_title = ''
    ticket_message_list = []
    unread_count_ticket = 0

    try:

        ticket_id = 0
        ticket_id_string = '0'
        try:
            ticket_id_string = request.path.split('/ticket_')[-1]
        except Exception as e:
            pass

        ticket_id = uuid.UUID(ticket_id_string)
        ticket_selected = Ticket.objects.get(id= ticket_id)
        ticket_title = ticket_selected.title
        
        # update seen status admin of ticket
        ticket_selected.admin_seen_status = Seen_Status.SEEN.value
        ticket_selected.save()

        ticket_message_list, total_count = get_ticket_messages_list(ticket_id)

        row_list_ticket, total_count_ticket, unread_count_ticket = get_ticket_list_page(request, 0, ROW_LIST_SHOW_COUNT, None)

        # update messsage seen status for client messages
        ticket_message_objects_list = Ticket_Message.objects.all().filter(ticket_id=ticket_id)
        for message in ticket_message_objects_list:
            if (message.user_type == User_Type.CLIENT.value):
                try:
                    message.seen_status = Seen_Status.SEEN.value
                    message.save()
                except Exception as e:
                    print(e)

    except Exception as e:
        print(e)
        pass

    return render(request, "home/ticket_messages_admin.html", {"msg": 'SUCCESS',
                                                                "segment": 'tickets_client', 
                                                                "title": ticket_title,
                                                                "ticket_id" : ticket_id_string,
                                                                "unread_count_ticket" : unread_count_ticket,
                                                                "ticket_message_list": ticket_message_list})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def new_message_admin(request):

    ticket_id = 0
    try:

        ticket_id = 0
        ticket_id_string = '0'
        try:
            ticket_id_string = request.path.split('/ticket_')[-1]
        except Exception as e:
            pass

        ticket_id = uuid.UUID(ticket_id_string)
        ticket_selected = Ticket.objects.get(id= ticket_id)

        message_body= ''

        if request.method == "POST": 
            message_body = request.POST.get('message_body').strip()

        # insert new message to ticket selected
        insert_message_db(request.user, message_body, ticket_selected, User_Type.ADMINISTRATOR.value)
        
        # update ticket message count and empty status
        try:
            ticket_selected.message_count = ticket_selected.message_count + 1
            ticket_selected.empty_status = Empty_Status.NOT_EMPTY.value
            ticket_selected.client_seen_status = Seen_Status.UNSEEN.value
            ticket_selected.answer_time_slot = datetime.datetime.now()

            if (ticket_selected.status == Ticket_Status.USER_ANSWERED.value or ticket_selected.status == Ticket_Status.USER_CREATED.value):
                ticket_selected.status = Ticket_Status.ADMIN_ANSWERED.value

            ticket_selected.save()
            
            # send email to client
            send_email_from_admin_to_client(ticket_selected.user, message_body)
        except Exception as e:
            print(colored(str(e), 'red'))



    except Exception as e:
        print(e)
        pass

    return redirect ('/ticket_messages_admin/ticket_' + str(ticket_id_string))

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def delete_message_admin(request):

    ticket_id = 0
    try:

        message_id = 0
        message_id_string = '0'
        ticket_id = ''

        try:
            message_id_string = request.path.split('/message_')[-1]
        except Exception as e:
            pass

        message_id = uuid.UUID(message_id_string)
        message_selected = Ticket_Message.objects.get(id= message_id)
        ticket_selected = message_selected.ticket
        ticket_id = ticket_selected.id

        # deleting messages from Ticket_Message first

        try:
            message_selected.delete()
        except Exception as e:
            print(colored('Exception in deleting message : ' + str(e), 'red'))

        # then update the ticket object Empty_Status and message_count and status
        try:
            if (ticket_selected.message_count == 1):
                ticket_selected.empty_status = Empty_Status.EMPTY.value

            ticket_selected.message_count = ticket_selected.message_count - 1
            ticket_selected.save()

        except Exception as e:
                print(colored('Exception in updating ticket : ' + str(e), 'red'))



    except Exception as e:
        print(e)
        pass

    return redirect ('/ticket_messages_admin/ticket_' + str(ticket_id))

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def new_ticket_admin(request):

    ticket_id = 0
    user_selected= None

    try:
        ticket_title= 'Not Inserted'

        if request.method == "POST": 
            ticket_title = request.POST.get('ticket_title').strip()
            user_email = request.POST.get('user_selected').strip()

        user_selected = User.objects.all().filter(email=user_email)[0]
        ticket_id = insert_ticket_db_admin(user_selected, ticket_title)
        

    except Exception as e:
        print(e)
        pass

    return redirect ('/ticket_messages_admin/ticket_' + str(ticket_id))

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def blog_admin(request):

    try:

        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_post, total_count_post = get_post_list_page(request, start, ROW_LIST_SHOW_COUNT, None)

        last_pagination = 1
        if (total_count_post == 0):
            last_pagination = 1
        elif (total_count_post % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_post / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_post / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/blog_admin.html", {"msg": 'SUCCESS',
                                                                "segment": 'blog_admin', 
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_post),
                                                                "total_items": total_count_post,
                                                                "post_list": row_list_post})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def new_post_admin(request):

    post_id = 0
    user_selected= None

    try:
        post_title= 'Not Inserted'
        post_image_name= 'Not Inserted'

        if request.method == "POST": 
            post_title = request.POST.get('post_title').strip()
            post_image_name = request.POST.get('post_image_name').strip()

        post_id = insert_post_db_admin(post_title, post_image_name)
        

    except Exception as e:
        print(e)
        pass

    return redirect ('/post_paragraph_admin/post_' + str(post_id))

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def new_discount_admin(request):

    discount_id = 0
    user_selected= None

    try:
        discount_code= 'Not Inserted'
        discount_value= 'Not Inserted'

        if request.method == "POST": 
            discount_code = request.POST.get('discount_code').strip()
            discount_value = request.POST.get('discount_value').strip()

        if (discount_code != 'Not Inserted' and discount_value != 'Not Inserted'):
            discount_id = insert_discount_db_admin(discount_code, discount_value)
        

    except Exception as e:
        print(e)
        pass

    return redirect ('/discounts_admin/')

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def delete_post(request):

    post_id = 0
    try:

        post_id = 0
        post_id_string = '0'
        try:
            post_id_string = request.path.split('/post_')[-1]
        except Exception as e:
            pass

        post_id = uuid.UUID(post_id_string)
        post_selected = Post.objects.get(id= post_id)

        post_paragraph_list = Post_Paragraph.objects.all().filter(post_id=post_id)

        # deleting all paragraph of post first
        for paragraph in post_paragraph_list:

            try:
                paragraph.delete()
            except Exception as e:
                print(colored('Exception in deleting paragraph : ' + str(e), 'red'))

        try:
            # then delete the post object itself
            post_selected.delete()
        except Exception as e:
                print(colored('Exception in deleting post : ' + str(e), 'red'))

    except Exception as e:
        print(e)
        pass

    return redirect ('/blog_admin/')

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def delete_discount(request):

    discount_id = 0
    try:

        discount_id = 0
        discount_id_string = '0'
        try:
            discount_id_string = request.path.split('/discount_')[-1]
        except Exception as e:
            pass

        discount_id = uuid.UUID(discount_id_string)
        discount_selected = Discount_Code.objects.get(id= discount_id)

        
        try:
            # then delete the discount object itself
            discount_selected.delete()
        except Exception as e:
                print(colored('Exception in deleting discount : ' + str(e), 'red'))

    except Exception as e:
        print(e)
        pass

    return redirect ('/discounts_admin/')

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def post_paragraph_admin(request):

    post_title = ''
    post_paragraph_list = []
    paragraph_count = 0

    try:

        post_id = 0
        post_id_string = '0'
        try:
            post_id_string = request.path.split('/post_')[-1]
        except Exception as e:
            pass

        post_id = uuid.UUID(post_id_string)
        post_selected = Post.objects.get(id= post_id)
        post_title = post_selected.title
        post_image_name = post_selected.image_name
    

        post_paragraph_list, total_count = get_post_paragraph_list(post_id)


    except Exception as e:
        print(e)
        pass

    return render(request, "home/post_paragraph_admin.html", {"msg": 'SUCCESS',
                                                                "segment": 'blog_admint', 
                                                                "title": post_title,
                                                                "image_name": post_image_name,
                                                                "post_id" : post_id_string,
                                                                "post_paragraph_list": post_paragraph_list})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def new_paragraph_admin(request):

    post_id = 0
    try:

        post_id = 0
        post_id_string = '0'
        try:
            post_id_string = request.path.split('/post_')[-1]
        except Exception as e:
            pass

        post_id = uuid.UUID(post_id_string)
        post_selected = Post.objects.get(id= post_id)

        paragraph_body= ''

        if request.method == "POST": 
            paragraph_body = request.POST.get('paragraph_body').strip()

        # insert new paragraph to post selected
        insert_paragraph_db(paragraph_body, post_selected)
        
    
    except Exception as e:
        print(colored(str(e), 'red'))



    except Exception as e:
        print(e)
        pass

    return redirect ('/post_paragraph_admin/post_' + str(post_id_string))

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def toggle_publish_status(request):

    post_id = 0
    try:

        post_id = 0
        post_id_string = '0'
        try:
            post_id_string = request.path.split('/post_')[-1]
        except Exception as e:
            pass

        post_id = uuid.UUID(post_id_string)
        post_selected = Post.objects.get(id= post_id)

        post_paragraph_list = Post_Paragraph.objects.all().filter(post_id=post_id)

        # toggle post_status publish

        if (post_selected.status == Post_Status.PUBLISHED.value):
            post_selected.status = Post_Status.UNPUBLISHED.value
        elif (post_selected.status == Post_Status.UNPUBLISHED.value):
            post_selected.status = Post_Status.PUBLISHED.value

        try:
            # then save the post object itself
            post_selected.save()
        except Exception as e:
                print(colored('Exception in updating post status : ' + str(e), 'red'))

    except Exception as e:
        print(e)
        pass

    return redirect ('/blog_admin/')

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def toggle_publish_status_discount(request):

    discount_id = 0
    try:

        discount_id = 0
        discount_id_string = '0'
        try:
            discount_id_string = request.path.split('/discount_')[-1]
        except Exception as e:
            pass

        discount_id = uuid.UUID(discount_id_string)
        discount_selected = Discount_Code.objects.get(id= discount_id)

        
        # toggle post_status publish
        if (discount_selected.status == Plan_Status.ENABLE.value):
            discount_selected.status = Plan_Status.DISABLE.value
        elif (discount_selected.status == Plan_Status.DISABLE.value):
            discount_selected.status = Plan_Status.ENABLE.value

        try:
            # then save the discount_code object itself
            discount_selected.save()
        except Exception as e:
                print(colored('Exception in updating discount status : ' + str(e), 'red'))

    except Exception as e:
        print(e)
        pass

    return redirect ('/discounts_admin/')

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def delete_paragraph_admin(request):

    post_id = 0
    try:

        paragraph_id = 0
        paragraph_id_string = '0'
        post_id = ''

        try:
            paragraph_id_string = request.path.split('/paragraph_')[-1]
        except Exception as e:
            pass

        paragraph_id = uuid.UUID(paragraph_id_string)
        paragraph_selected = Post_Paragraph.objects.get(id= paragraph_id)
        post_selected = paragraph_selected.post
        post_id = post_selected.id

        # deleting paragraph from Post_Paragraph first

        try:
            paragraph_selected.delete()
        except Exception as e:
            print(colored('Exception in deleting paragraph : ' + str(e), 'red'))

        
        except Exception as e:
                print(colored('Exception in updating post : ' + str(e), 'red'))



    except Exception as e:
        print(e)
        pass

    return redirect ('/post_paragraph_admin/post_' + str(post_id))

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def edit_paragraph_admin(request):

    post_id = 0

    try:

        paragraph_id = 0
        paragraph_id_string = '0'
        post_id = ''

        try:
            paragraph_id_string = request.path.split('/paragraph_')[-1]
        except Exception as e:
            pass

        paragraph_id = uuid.UUID(paragraph_id_string)
        paragraph_selected = Post_Paragraph.objects.get(id= paragraph_id)
        post_selected = paragraph_selected.post
        post_id = post_selected.id

        # edit paragraph from Post_Paragraph first
        try:
            paragraph_body= ''
            if request.method == "POST": 
                paragraph_body = request.POST.get('modal_paragraph_edit').strip()
                paragraph_selected.body = paragraph_body
                paragraph_selected.save()

        except Exception as e:
            print(colored('Exception in editing paragraph : ' + str(e), 'red'))

        
        except Exception as e:
                print(colored('Exception in updating post : ' + str(e), 'red'))



    except Exception as e:
        print(e)
        pass

    return redirect ('/post_paragraph_admin/post_' + str(post_id))

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def edit_post_title_admin(request):

    try:

        post_id = 0
        post_id_string = '0'

        try:
            post_id_string = request.path.split('/post_')[-1]
        except Exception as e:
            pass

        post_id = uuid.UUID(post_id_string)
        post_selected = Post.objects.get(id= post_id)

        # edit post title from Post
        try:
            title= ''
            if request.method == "POST":
               
                title = request.POST.get('modal_title_edit').strip()
                
                post_selected.title = title
                post_selected.save()

        except Exception as e:
            print(colored('Exception in editing post title : ' + str(e), 'red'))

        
    except Exception as e:
        print(colored('Exception in updating post title : ' + str(e), 'red'))

    return redirect ('/post_paragraph_admin/post_' + str(post_id))

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def edit_post_image_name_admin(request):

    try:

        post_id = 0
        post_id_string = '0'

        try:
            post_id_string = request.path.split('/post_')[-1]
        except Exception as e:
            pass

        post_id = uuid.UUID(post_id_string)
        post_selected = Post.objects.get(id= post_id)

        # edit post title from Post
        try:
            image_name= ''
            if request.method == "POST": 
                image_name = request.POST.get('modal_image_name_edit').strip()
                
                post_selected.image_name = image_name
                post_selected.save()

        except Exception as e:
            print(colored('Exception in editing post image_name : ' + str(e), 'red'))

        
    except Exception as e:
        print(colored('Exception in updating post image_name : ' + str(e), 'red'))

    return redirect ('/post_paragraph_admin/post_' + str(post_id))

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def update_bulk_status_ajax_admin(request):
    current_pagination = 1   
    try:
        current_pagination = int (request.path.split('/P')[-1])
    except Exception as e:
        pass
    start = (current_pagination-1) * ROW_LIST_SHOW_COUNT
    row_list_bulk, total_count_bulk = get_bulk_list_page_mongodb(request, start, ROW_LIST_SHOW_COUNT, None)

    if request.is_ajax and request.method == "GET":
        return render(request, 'includes/bulk_status_table_admin.html', {'bulk_list':row_list_bulk})

@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def upload_post_image(request):

    try:

        post_id = 0
        post_id_string = '0'

        try:
            post_id_string = request.path.split('/post_')[-1]
        except Exception as e:
            pass

        post_id = uuid.UUID(post_id_string)
        post_selected = Post.objects.get(id= post_id)

        try:
            image_name= ''
            if request.method == "POST": 

                # edit post image name
                image_name = request.POST.get('input_post_image').strip()
                post_selected.image_name = image_name
                post_selected.save()

                # transfer image to directory
                form = FileForm(request.POST, request.FILES)
                if form.is_valid():
                    print('FFFFFFFFFFFFFFFFFF')
                    
                    print(type(request.FILES["input_post_image"]))
        except Exception as e:
            print(colored('Exception in editing post image_name : ' + str(e), 'red'))

        
    except Exception as e:
        print(colored('Exception in updating post image_name : ' + str(e), 'red'))

    return redirect ('/post_paragraph_admin/post_' + str(post_id))


@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def discounts_admin(request):

    try:

        current_pagination = 1    
        try:
            current_pagination = int (request.path.split('/P')[-1])
        except Exception as e:
            pass
        
        start = (current_pagination-1) * ROW_LIST_SHOW_COUNT

        row_list_discount, total_count_discount = get_discount_list_page(request, start, ROW_LIST_SHOW_COUNT, None)

        last_pagination = 1
        if (total_count_discount == 0):
            last_pagination = 1
        elif (total_count_discount % ROW_LIST_SHOW_COUNT == 0):
            last_pagination = int (total_count_discount / ROW_LIST_SHOW_COUNT)
        else:
            last_pagination = int (total_count_discount / ROW_LIST_SHOW_COUNT) + 1

    except Exception as e:
        print(e)
        pass

    return render(request, "home/discount_admin.html", {"msg": 'SUCCESS',
                                                                "segment": 'purchase-discounts_admin', 
                                                                "current_pagination": current_pagination,
                                                                "last_pagination": last_pagination,
                                                                "page_limit": len(row_list_discount),
                                                                "total_items": total_count_discount,
                                                                "discount_list": row_list_discount})


@staff_member_required(login_url="/login/")
@login_required(login_url="/login/")
def delete_user_admin(request):

    user_id = 0
    try:

        user_id = 0
        user_id_string = '0'
        try:
            user_id_string = request.path.split('/user_')[-1]
        except Exception as e:
            pass

        user_id = int(user_id_string)
        user_selected = User.objects.get(id= user_id)
        
        # deleting user from Ticket_Message
        try:
            Ticket_Message_selected = Ticket_Message.objects.all().filter(user_id= user_id)
            for item in Ticket_Message_selected:
                item.delete()
        except Exception as e:
            print(colored('Exception in deleting Ticket_Message : ' + str(e), 'red'))


        # deleting user from User_Log
        try:
            query_log = {'user': user_id_string}
            collection_User_Log.delete_many(query_log)
            
        except Exception as e:
            print(colored('Exception in deleting User_Log : ' + str(e), 'red'))

        # deleting user from Ticket
        try:
            Ticket_selected = Ticket.objects.all().filter(user_id= user_id)
            for item in Ticket_selected:
                
                Ticket_message_selected = Ticket_Message.objects.all().filter(ticket_id= item.id)
                for message in Ticket_message_selected:
                    message.delete()

                item.delete()
        except Exception as e:
            print(colored('Exception in deleting Ticket : ' + str(e), 'red'))

        # deleting user from Payment
        try:
            Payment_selected = Payment.objects.all().filter(user_id= user_id)
            for item in Payment_selected:
                item.delete()
        except Exception as e:
            print(colored('Exception in deleting Payment : ' + str(e), 'red'))


        # deleting user from Bulk and Requests mongo
        try:
            query_request = {'user': user_id_string}
            collection_Requests.delete_many(query_request)
            collection_Bulks.delete_many(query_request)

        except Exception as e:
            print(colored('Exception in Bulk and Requests mongo : ' + str(e), 'red'))

        # Finally deleting user from User and User_Other_Fields
        try:
            user_selected.delete()
            User_Other_Fields_selected = User_Other_Fields.objects.all().filter(user_id= user_id)
            for item in User_Other_Fields_selected:
                item.delete()
        except Exception as e:
            print(colored('Exception in deleting User : ' + str(e), 'red'))


    except Exception as e:
        print(e)
        pass

    return redirect ('/userslist_admin/')
